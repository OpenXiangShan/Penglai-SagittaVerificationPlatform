#!/usr/bin/python
# coding=utf-8
def install(package_name):
    try:
        __import__(package_name)
    except ImportError:
        print('installing {}'.format(package_name))
        if hasattr(pip, 'main'):
            pip.main(['install', package_name])
        else:
            pip._internal.main(['install', package_name])

import os
import sys
import time
import re
import random
import pip
import datetime
install("pathlib")
install("xlrd")
install("xlwt")
install("xlutils")
install("matplotlib")
install("openpyxl")
install("csv")
install("numpy")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import multiprocessing
import subprocess
import ast
import shlex
import pathlib
import xlrd
import xlwt
import urllib
import urllib.parse
import base64
from xlutils.copy import copy as xlcopy
import sqlite3
import csv, copy
import openpyxl
from openpyxl.styles import  PatternFill
import numpy as np
from crontab import CronTab
try:
    # Python3
    install("configparser")
    install("paramiko")
    install("hashlib")
    install("requests")
    install("json")
    from configparser import ConfigParser
    # send info lib
    import paramiko
    import hashlib
    import requests
    from requests import Request,Session
    import json
except ImportError:
    print("Import Error!!")

# global value
rerunTimes = 0


class SqlMethod:
    DBname = ""  #数据库名称

    def __init__(self,DBname):
        self.DBname = DBname
        self.con = sqlite3.connect('{}.db'.format(self.DBname))
        self.cur = self.con.cursor()

    def __del__(self):
        #最后要记得断开连接
        self.release()

    def release(self):
        self.cur.close()
        self.con.close()

    def BuildNewTable(self,TableName):
        #Table每一列的信息，可以修改成自己需要的格式，这里建议ID只能递增1，毕竟是PRIMARY KEY，变化规则要确定好
        self.cur.execute('''
                CREATE TABLE IF NOT EXISTS {}(
                ID                 INT           PRIMARY KEY,
                DATE               NONE,
                TOTAL              INT,
                EXECUTED           INT,
                FATAL              INT,
                ERROR              INT,
                WARNING            INT,
                RUNNING            INT,
                NOT_START          INT,
                PASS               INT,
                PASSRate           TEXT,
                code_coverage      TEXT,
                LINE               TEXT,
                COND               TEXT,
                TOGL               TEXT,
                FSM                TEXT,
                func_couverage     TEXT,
                Tapd_closed        INT,
                Tapd_not_closed    INT,
                Plan_coverage      TEXT
                )
                '''.format(TableName))

    def GetTableName(self):
        #获得Table的名称
        self.cur.execute("SELECT name FROM sqlite_master WHERE type=\'table\'")
        tmpTables = self.cur.fetchall()
        Tables = []
        for ele in tmpTables:
            Tables.append(ele[0])
        return Tables

    def GetTableTitle(self):
        #获得Table每一列的信息内容，默认数据库中所有表的格式一致
        TableNames = self.GetTableName()
        self.cur.execute('SELECT * FROM \'{}\' '.format(str(TableNames[0])))
        ColNameList = [tuple[0] for tuple in self.cur.description]
        return ColNameList

    def GetValueNum(self,TableName):
        #获得Table中的数据数量
        self.cur.execute('SELECT COUNT(ID) FROM \'{}\' '.format(str(TableName)))
        return self.cur.fetchall()[0][0]

    def GetLastId(self,TableName):
        #获得Table里最新的数据
        if TableName not in self.GetTableName():
            print("No {} in SQL, please check! {}".format(TableName,self.GetTableName))
            return
        else:
            if self.GetValueNum(TableName) == 0:
                return 0
            else:
                self.cur.execute('SELECT MAX(ID) FROM \'{}\''.format(str(TableName)))
            return self.cur.fetchall()[0][0]

    def InsertValue(self,TableName,Values):
        #对应的Table插入数据
        TableTitles = self.GetTableTitle()
        if len(TableTitles) != len(Values):
            #检查插入数据数量对不对
            print("Table Tiles is not equal to values: \n Title: {} \n Value: {}".format(TableTitles,Values))
        else:
            if TableName not in self.GetTableName():
                #检查Table的名字对不对
                print("No {} in SQL, please check! {}".format(TableName,self.GetTableName))
            else:
                tmpStr = "\'"
                tmpStr += "\',\'".join([str(i) for i in Values])
                tmpStr += "\'"
                #sqlcmd = "INSERT INTO {} ({}) VALUES({})".format(TableName,",".join(TableTitles), tmpStr)
                sqlcmd = "INSERT INTO {} VALUES({})".format(TableName, tmpStr)
                self.cur.execute(sqlcmd)
                self.con.commit()

    def GetLastVal(self,TableName):
        #获得Table内最新的数据
        LastValue = ()
        if TableName not in self.GetTableName():
            print("No {} in SQL, please check! {}".format(TableName,self.GetTableName))
        else:
            self.cur.execute('SELECT * FROM \'{}\' WHERE ID = (SELECT MAX(ID) FROM \'{}\')'.format(TableName,TableName))
            LastValue = self.cur.fetchall()
        return LastValue[0]


class XrunTestPointReverse:
    def ReadCsvFile(self, ModeName):
        CsvTable = []
        FileName = ModeName + '_all_bins.csv' if ModeName != '' else 'all_bins.csv'
        CsvFile = open(FileName,'r')
        CsvReader = csv.reader(CsvFile)
        CsvTableTile = next(CsvReader)
        for row in CsvReader:
            CsvTable.append(row)
        CsvFile.close()
        return CsvTableTile, CsvTable

    def FcovBelongJudge(self, tmpIdList, FcovIdList):
        for i in range(len(FcovIdList)):
            if tmpIdList[i] != FcovIdList[i]:
                return False
        return True

    def FcovLv(self, tmpId, FcovIdList):
        # TODO Number Id length cannot be determined
        tmpId = [int(i) for i in tmpId.split('.')]
        if len(tmpId) < len(FcovIdList):  # level > fcov
            FcovBelong = False
            FcovLevel = '' # ["",total,cg,cp,bins]
        elif len(tmpId) == len(FcovIdList) and self.FcovBelongJudge(tmpId,FcovIdList):  # level == total
            FcovBelong = True
            FcovLevel  = 'total'
        elif len(tmpId) == len(FcovIdList) + 2 and self.FcovBelongJudge(tmpId,FcovIdList):  # level == cg
            FcovBelong = True
            FcovLevel  = 'cg'
        elif len(tmpId) == len(FcovIdList) + 3 and self.FcovBelongJudge(tmpId,FcovIdList):  # level == cp
            FcovBelong = True
            FcovLevel  = 'cp'
        elif len(tmpId) == len(FcovIdList) + 4 and self.FcovBelongJudge(tmpId,FcovIdList):  # level == cg
            FcovBelong = True
            FcovLevel  = 'bins'
        else:
            FcovBelong = False
            FcovLevel = ''
        return FcovBelong, FcovLevel

    def InsertBinsInFcovDict(self, BinsInfo, FcovDict):
        for i in range(len(FcovDict['cp_list'])):
            if BinsInfo['root'] == FcovDict['cp_list'][i]['Name']:
                FcovDict['cp_list'][i]['leaf'].append(BinsInfo['Name'])
                FcovDict['bins_list'].append(BinsInfo)
                break
        return FcovDict

    def CumulateSameBins(self, OldFcovDict):
        NewFcovDict = {'total':[], 'cg_list':[],'cp_list':[],'bins_list':[]}
        NewFcovDict['total'].append(OldFcovDict['total'][0])
        for ele in OldFcovDict['cg_list']:
            NewFcovDict['cg_list'].append(ele)
        for ele in OldFcovDict['cp_list']:
            ele['leaf'].clear()
            NewFcovDict['cp_list'].append(ele)
        CumBins = {'Tag':'bins','Name':"",'Percent':"",'root':"",'leaf':[]}
        CumResult = 0
        CumNum = 0
        for BinsInfo in OldFcovDict['bins_list']:
            if '[' in BinsInfo['Name']:
                tmpName = BinsInfo['Name'].split('[')[0]
                if tmpName != CumBins['Name'] or BinsInfo['root'] != CumBins['root']: # new list of bins
                    # deal old bins
                    CumBins['Percent'] = f2p(CumResult/CumNum if CumNum!=0 else 0)
                    self.InsertBinsInFcovDict(copy.copy(CumBins),NewFcovDict)
                    # initial new bins
                    CumBins['Name'] = tmpName
                    CumBins['Percent'] = ''
                    CumBins['root'] = BinsInfo['root']
                    CumResult = p2f(BinsInfo['Percent'])
                    CumNum = 1
                else:
                    CumResult += p2f(BinsInfo['Percent'])
                    CumNum += 1
            else:
                NewFcovDict = self.InsertBinsInFcovDict(BinsInfo,NewFcovDict)
        CumBins['Percent'] = f2p(CumResult/CumNum)
        self.InsertBinsInFcovDict(copy.copy(CumBins),NewFcovDict)
        return NewFcovDict

    def GetFcovDict(self,TableTitleList, Table):
        FcovDict = {'total':[], 'cg_list':[],'cp_list':[],'bins_list':[]}
        FcovId = ""
        FcovTarget = "$unit"
        for ele in Table:
            if FcovTarget in ele[3]:
                    FcovId = ele[0]
        FcovIdList = [int(ele) for ele in FcovId.split('.')] # 3 valid Id num like 1.2.17
        tmpCg = ''
        tmpCp = ''
        for ele in Table:
            FcovBelong, FcovLevel = self.FcovLv(ele[0],FcovIdList)
            if FcovBelong:
                tmpDict = {'Tag':FcovLevel,'Name':ele[3],'Percent':ele[4],'root':"",'leaf':[]} # Tree link, [] is none
                if FcovLevel == 'total':# total info
                    FcovDict['total'].append(tmpDict)
                elif FcovLevel == 'cg':
                    tmpDict['root'] = FcovDict['total'][0]['Name']
                    FcovDict['total'][0]['leaf'].append(tmpDict['Name']) # update total leaf
                    FcovDict['cg_list'].append(tmpDict)
                    tmpCg = tmpDict['Name']
                elif FcovLevel == 'cp':
                    tmpDict['root'] = tmpCg
                    FcovDict['cp_list'].append(tmpDict)
                    for tmp in FcovDict['cg_list']:
                        if tmp['Name'] == tmpCg:
                            tmp['leaf'].append(tmpDict['Name']) # update root's leaf
                            break
                    tmpCp = tmpDict['Name']
                elif FcovLevel == 'bins':
                    tmpDict['root'] = tmpCp
                    FcovDict['bins_list'].append(tmpDict)
                    for tmp in FcovDict['cp_list']:
                        if tmp['Name'] == tmpCp:
                            tmp['leaf'].append(tmpDict['Name']) # update root's leaf
                            break
                else:
                    print('Warning! Unknown Type {}'.format(FcovLevel))
        return FcovDict

    def Mode2FcovDict(self, ModeName):
        TableTitleList, Table = self.ReadCsvFile(ModeName)
        FcovDict = self.GetFcovDict(TableTitleList, Table )
        FcovDict = self.CumulateSameBins(FcovDict)
        return FcovDict


class anti_annotation:
    def __init__(self, FcovDict, FilePath, XLName):
        self.FcovDict = FcovDict
        self.FilePath = FilePath
        self.XLName = XLName

    def CheckFileExists(self):
        if not os.path.exists(os.path.join(self.FilePath,self.XLName)):
            print('TestPoint File doesn\'t exist. Please, check file ( {} )'.format(os.path.join(self.FilePath,self.XLName)))
            sys.exit()
        else:
            return

    def CheckCgInTotal(self,CgName):
        if CgName not in self.FcovDict['total'][0]['leaf']:
            print('Warning! No Cg named ({}) in fcov'.format(CgName))
            return None
        else:
            return self.FcovDict['total'][0]['leaf'].index(CgName)

    def CheckCpInCg(self,CpName,CgIdx):
        if CpName not in self.FcovDict['cg_list'][CgIdx]['leaf']:
            print('Warning! No Cp named ({}) in Cg({})'.format(CpName,self.FcovDict['cg_list'][CgIdx]['Name']))
            return None
        else:
            return [ele['Name'] for ele in self.FcovDict['cp_list']].index(CpName)

    def CheckBinsInCp(self,BinsName,CpIdx):
        if BinsName not in self.FcovDict['cp_list'][CpIdx]['leaf']:
            print('Warning! No Bins named ({}) in Cp({})'.format(BinsName,self.FcovDict['cp_list'][CpIdx]['Name']))
            return None
        else:
            Idx = 0
            for ele in self.FcovDict['bins_list']:
                if ele['Name'] == BinsName and ele['root'] == self.FcovDict['cp_list'][CpIdx]['Name']:
                    return Idx
                else:
                    Idx += 1

    def CovPath2CovPercent(self, CovPath):
        CovPercent = ''
        CovPath = CovPath.replace(' ','')
        CovPath = CovPath.split('::')[-1]
        if CovPath == '':
            CovPercent = ''
        else:
            # CovPath format: uint::fcov::xxx_cg.xxx_cp.xxx '\n'  unit::fcov::xxx_cg.xxx_cp.xxx
            CovPathList = CovPath.split('.')
            if len(CovPathList) == 1: # cg
                CgName = CovPathList[0]
                CovPercent = self.FcovDict['cg_list'][self.CheckCgInTotal(CgName)]['Percent'] if self.CheckCgInTotal(CgName) else ""
            elif len(CovPathList) == 2: # cp
                CgName = CovPathList[0]
                CpName = CovPathList[1]
                if self.CheckCgInTotal(CgName) is not None:
                    CgIdx = self.CheckCgInTotal(CgName)
                    CovPercent = self.FcovDict['cp_list'][self.CheckCpInCg(CpName,CgIdx)]['Percent'] if self.CheckCpInCg(CpName,CgIdx) else ""
            elif len(CovPathList) == 3: # bins
                CgName = CovPathList[0]
                CpName = CovPathList[1]
                BinsName = CovPathList[2]
                if self.CheckCgInTotal(CgName) is not None:
                    CgIdx = self.CheckCgInTotal(CgName)
                    if self.CheckCpInCg(CpName,CgIdx) is not None:
                        CpIdx = self.CheckCpInCg(CpName,CgIdx)
                        CovPercent = self.FcovDict['bins_list'][self.CheckBinsInCp(BinsName,CpIdx)]['Percent'] if self.CheckBinsInCp(BinsName,CpIdx) else ""
            else:
                print('Unknown Type ( {} )'.format(CovPath))
        return CovPercent

    def p2f(self,x):
        return float(x.strip('%'))/100

    def FillColor(self, Cell, ColorName):
        if ColorName == 'red':
            Color = 'F05D43'
        if ColorName == 'green':
            Color = '78EA5C'
        if ColorName == 'yellow':
            Color = 'E4F125'
        if ColorName == 'gray':
            Color = '808B96'
        fill = PatternFill("solid", fgColor=Color)
        Cell.fill = fill
        return Cell

    def WriteAntiAnnotation(self):
        TotalPercentList = []
        self.CheckFileExists()
        CovPathTitle = 'reverse_testpoint'
        WeightTitle = 'weight'
        AntiAnnotationTitle = 'reverse_percent'
        workbook = openpyxl.load_workbook(os.path.join(self.FilePath,self.XLName))
        worksheet = workbook.worksheets[-1] # use last sheet
        TitleList = [cell.value for cell in list(worksheet.rows)[0]]
        CovPathTitleIdx = TitleList.index(CovPathTitle) + 1
        WeightTitleIdx = TitleList.index(WeightTitle) + 1
        AntiAnnotationTitleIdx = TitleList.index(AntiAnnotationTitle) + 1
        MaxRow = worksheet.max_row
        for i in range(2,MaxRow+1):
            CellValue = worksheet.cell(row = i, column = CovPathTitleIdx).value  # get value path
            if CellValue:
                FcovPathList = CellValue.split('\n')
                FcovPercent = []
                for FcovPath in FcovPathList:
                    tmpResult = self.CovPath2CovPercent(FcovPath)
                    FcovPercent.append(tmpResult if tmpResult != '' else "None")
                    if tmpResult != '':
                        TotalPercentList.append(tmpResult)
                    else:
                        TotalPercentList.append('0%')
                    tmpCell = worksheet.cell(i,AntiAnnotationTitleIdx,'\n'.join(FcovPercent))
                    if "None" in FcovPercent:
                        tmpCell = self.FillColor(tmpCell,'gray')
                    else:
                        FloatFcovPercent = [self.p2f(ele) for ele in FcovPercent]
                        FloatFcovMean = np.mean(FloatFcovPercent)
                        if FloatFcovMean < 0.9:
                            tmpCell = self.FillColor(tmpCell,'red')
                        elif FloatFcovMean >= 0.9 and FloatFcovMean < 1:
                            tmpCell = self.FillColor(tmpCell,'yellow')
                        elif FloatFcovMean == 1:
                            tmpCell = self.FillColor(tmpCell,'green')
            else:
                continue
        workbook.save(os.path.join(self.FilePath,self.XLName))
        return np.mean([self.p2f(ele) for ele in TotalPercentList])


class CrontabMethod:

    def __init__(self,cmt):
        self.MyCron = CronTab(user=True)
        self.comment = cmt

    def PrintAllJob(self):
        for job in self.MyCron:
            print(job)

    def FindJobInTab(self,Cmd):
        for job in self.MyCron:
            if job.command in Cmd:
                return True, job
        return False, None

    def FindJobByCmmt(self,Cmmt):
         for job in self.MyCron:
            if Cmmt == job.comment:
                return True, job
         return False, None

    def SetCrontab(self, Cmd, SetTime):
        tmpFindResult, tmpJob = self.FindJobInTab(Cmd)
        if tmpFindResult: # update
            tmpJob.setall(SetTime)
            self.MyCron.write()
        else: # create
            job = self.MyCron.new(command=Cmd, comment=self.comment)
            job.setall(SetTime)
            self.MyCron.write()
        print("SetCronTab Success")

    def SetCronTabByDeltaMinutes(self,Cmd,MinuteTime):
        job = self.MyCron.new(command=Cmd,comment=self.comment)
        job.minute.every(MinuteTime)
        self.MyCron.write()
        print("SetCronTab Success")

    def SetCronTabByDeltaDays(self,Cmd,DayTime):
        tmpFindResult, tmpJob = self.FindJobInTab(Cmd)
        if tmpFindResult:
            tmpJob.day.every(DayTime)
            self.MyCron.write()
        else:
            job = self.MyCron.new(command=Cmd,comment=self.comment)
            job.day.every(DayTime)
            self.MyCron.write()
        print("SetCronTab Success")

    def SetCronTabByClock(self,Cmd,Hour,Minute):
        tmpFindResult, tmpJob = self.FindJobInTab(Cmd)
        if tmpFindResult:
            tmpJob.hour.on(Hour)
            tmpJob.minute.on(Minute)
            self.MyCron.write()
        else:
            job = self.MyCron.new(command=Cmd,comment=self.comment)
            job.hour.on(Hour)
            job.minute.also.on(Minute)
            self.MyCron.write()
        print("SetCronTab Success on  {}:{} o'clock".format(Hour,Minute))

    def DelCrontabByCmd(self,CmdKeyWord):
        for job in self.MyCron:
            if CmdKeyWord in job.command:
                self.MyCron.remove(job)
                self.MyCron.write()
        print("Delete Crontab Cmd Success")

    def DelCrontabByCmmt(self,CmmtKeyWord):
        for job in self.MyCron:
            if CmmtKeyWord in job.comment:
                self.MyCron.remove(job)
                self.MyCron.write()
        print("Delete Crontab Cmd Success")

    # for regress Do some change
    # generate local shell regress script
    def GenLocalShScript(self, LocalPath, RegrIniFileName, RegrShName, CshFile):
        ShInfo = "#!/bin/csh\n"
        ShInfo += "/usr/bin/xfce4-terminal -x csh -c "
        # source csh_common is add for Yinhe platform
        if CshFile != '':
            ShInfo += "\"source /sdata/home/data_exchange/shareDir/csh_common; cd {} ; source {}; ".format(LocalPath, CshFile) 
        else:
            ShInfo += "\"source /sdata/home/data_exchange/shareDir/csh_common; cd {} ; ".format(LocalPath)

        if RegrIniFileName == '':
            ShInfo += "make do_regr >& ./regress_`date+%m%d_%H%M`.log; ".format(LocalPath)
        else:
            ShInfo += "make do_regr regr_ini={} >& ./regress_`date+%m%d_%H%M`.log; ".format(RegrIniFileName)
        ShInfo += " exec csh\""
        tmpFile = open(os.path.abspath(os.path.join(LocalPath,'{}.sh'.format(RegrShName))),'w')
        tmpFile.write(ShInfo)
        tmpFile.close()
        print('Generate local Regress Shell Script Done!')

    # generate crontab regress command
    def GenRegressCmd(self, LocalPath, RegrShName):
        Cmd = " export DISPLAY=:1.0 && sh {} ".format(os.path.abspath(os.path.join(LocalPath,"{}.sh".format(RegrShName))))
        return Cmd


def SplitIgnorQuotes(s):
    tmp_List = re.split(r'([^ ]*"[^"]*")', s)
    sList = []
    newList = []
    tmpList = []
    for tmp in tmp_List:
        if re.search('"',tmp):
            sList.append(tmp)
        else:
            sList.extend(tmp.split())
    for ele in sList:
       if re.search("udf=",ele):
            tmpList = ele.replace('"','').replace('udf=','').split("+define+")
            for udfInfo in tmpList:
                udfInfo.replace(" ",'')
                if udfInfo != '':
                    newList.append('udf+="+define+'+udfInfo + '"')
       else:
            newList.append(ele)
    return newList


def CallOperate(Operate,ErrorInfo):
    print('CallOperate of ({0}) {1} {2}'.format(SplitIgnorQuotes(Operate), 'with type ',type(Operate)),flush = True)
    try:
        subprocess.check_call(SplitIgnorQuotes(Operate))
    except: #subprocess.CalledProcessError as ErrorInfo:
        print('Error:{}'.format(ErrorInfo))
        sys.exit()


def IniPathMatch(IniFileName):
    IniFilePathList = ['./','../script','../regress','./error_ini_dir']
    for tmpPath in IniFilePathList:
        tmpFile = os.path.abspath(os.path.join(tmpPath,IniFileName))
        if os.path.exists(tmpFile):
            return tmpFile
    print('Error! Script cannot find ini file!')
    sys.exit()
    return ''


def DoIniParser():
    print('''read ini file''',flush = True)
    try:
        IniFileName = sys.argv[1]
    except:
        IniFileName = "regress.ini"
    IniFile = IniPathMatch(IniFileName)
    Ini = ConfigParser()
    Ini.read(IniFile)
    print("IniFile is ",IniFile)
    GeneralDict = {}
    PreRegressDict = {}
    RegressDict = {}
    for option in Ini.options('GENERAL'):
        GeneralDict[option] = Ini.get('GENERAL',option)
    for option in Ini.options('PRE_REGRESSION'):
        PreRegressDict[option] = Ini.get('PRE_REGRESSION',option)
    for option in Ini.options('REGRESSION'):
        RegressDict[option] = Ini.get('REGRESSION',option)
    IniCheck(GeneralDict, PreRegressDict, RegressDict)
    return GeneralDict, PreRegressDict, RegressDict


def IniCheck(GeneralDict, PreRegressDict, RegressDict):
    ''' General Dict check'''
    DictCheck(GeneralDict,'tools'            ,'none')
    DictCheck(GeneralDict,'operation'        ,'none')
    DictCheck(GeneralDict,'url'              ,'')
    DictCheck(GeneralDict,'clone_workspace'  ,'')
    DictCheck(GeneralDict,'sim_path'         ,'')
    DictCheck(GeneralDict,'module_name'      ,'dut')
    DictCheck(GeneralDict,'excel_name'       ,'dut_excel')
    DictCheck(GeneralDict,'excel_path'       ,'./')
    DictCheck(GeneralDict,'sim_tools'        ,'xrun')
    DictCheck(GeneralDict,'paral_run_num'    ,'10')
    DictCheck(GeneralDict,'tapd_enable'      ,'disable')
    DictCheck(GeneralDict,'tapd_module_name' ,'')
    DictCheck(GeneralDict,'anti_annotation_enable'      ,'disable')
    DictCheck(GeneralDict,'testpoint_file_name'      ,'')
    DictCheck(GeneralDict,'fail_recmp'       ,'disable')
    DictCheck(GeneralDict,'recmp_opts'       ,'')
    DictCheck(GeneralDict,'fail_rerun'       ,'disable')
    DictCheck(GeneralDict,'rerun_opts'       ,'')
    DictCheck(GeneralDict,'ignore_warning_en'           ,'disable')
    DictCheck(GeneralDict,'fatal_key_words'             ,'[]')
    DictCheck(GeneralDict,'error_key_words'             ,'[]')
    DictCheck(GeneralDict,'warning_key_words'           ,'[]')
    DictCheck(GeneralDict,'ignore_fatal_key_words'      ,'{}')
    DictCheck(GeneralDict,'ignore_error_key_words'      ,'{}')
    DictCheck(GeneralDict,'ignore_warning_key_words'    ,'{}')
    DictCheck(GeneralDict,'total_ignore_log_list'       ,'[]')
    DictCheck(GeneralDict,'start_chk_fatal_key_word'    ,'')
    DictCheck(GeneralDict,'start_chk_error_key_word'    ,'')
    DictCheck(GeneralDict,'start_chk_warning_key_word'  ,'')
    DictCheck(GeneralDict,'send_to_workwx'              ,'disable')
    DictCheck(GeneralDict,'send_to_email'               ,'disable')
    DictCheck(GeneralDict,'receiver_list'               ,'[]')
    DictCheck(GeneralDict,'set_daily_regress'           ,'disable')
    DictCheck(GeneralDict,'regress_date_hour'           ,'0')
    DictCheck(GeneralDict,'regress_date_minute'         ,'0')
    DictCheck(GeneralDict,'project_source_file'         ,'')
    ''' Pre_regerssion Dict check'''
    DictCheck(PreRegressDict,'pre_script_enable'    ,'disable')
    DictCheck(PreRegressDict,'pre_make_enable'      ,'disable')
    DictCheck(PreRegressDict,'post_script_enable'   ,'disable')
    DictCheck(PreRegressDict,'post_make_enable'     ,'disable')
    DictCheck(PreRegressDict,'pre_script'           ,'')
    DictCheck(PreRegressDict,'pre_make'             ,'')
    DictCheck(PreRegressDict,'post_script'          ,'')
    DictCheck(PreRegressDict,'post_make'            ,'')
    ''' Regerssion Dict check'''
    DictCheck(RegressDict,'dotransport'    ,'disable')
    DictCheck(RegressDict,'transport_while_regress','disable')
    DictCheck(RegressDict,'docompile'      ,'enable')
    DictCheck(RegressDict,'fcov'           ,'on')
    DictCheck(RegressDict,'ccov'           ,'on')
    DictCheck(RegressDict,'fcov_key_word'  ,'_common_pkg')
    DictCheck(RegressDict,'udf'            ,'')
    DictCheck(RegressDict,'cmp_opts'       ,'')
    DictCheck(RegressDict,'dorotation'     ,'disable')
    DictCheck(RegressDict,'del_pass_log'     ,'enable')
    DictCheck(RegressDict,'del_err_cov'     ,'disable')


def DictCheck(Dict, keys, default_value):
    if keys not in Dict.keys():
        Dict[keys] = default_value
        print('ini file miss [{}] option, set value: {}'.format(keys,default_value))
    else:
        if '\t' in Dict[keys]:
            Dict[keys] = Dict[keys].replace('\t','')


def CoSvnUrl(url):
    tmp = re.sub("svn://.*:\d+/", "", url)
    tmp_list = tmp.split('/')
    if(len(tmp_list)==1):
        CallOperate("svn co {}".format(url),'Do svn co Failed!!!')
    else:
        for i in range(len(tmp_list)-1):
            if os.path.exists(tmp_list[i]) is False:
                os.makedirs(tmp_list[i])
            os.chdir(tmp_list[i])
        CallOperate("svn co {}".format(url),'Do svn co Failed!!!')


def DoSourcePrj(SimPath):
    os.chdir(SimPath)
    source_cmd = "source ../../../../scr/project.cshrc"
    split_flag = "regress.py source project.cshrc and get env"
    split_cmd = 'echo "{}"'.format(split_flag)
    env_cmd = "env"
    get_source_env = subprocess.check_output(["/bin/csh", "-c", source_cmd + " && " + split_cmd + " && " + env_cmd])
    source_env_list = "{}".format(get_source_env).split(split_flag)[-1][2:-1].split("\\n")[:-1]
    for source_env in source_env_list:
        tmp = source_env.split("=")
        os.environ[tmp[0]] = tmp[1]


def GenRegressPath(GeneralDict):
    global RegressTime
    print('''Get Path''')
    if GeneralDict['operation'] in ['clone','co']:
        RegressPath = os.path.abspath(os.path.join(GeneralDict['clone_workspace'],RegressTime))
        SimPath = os.path.abspath(os.path.join(RegressPath,GeneralDict['sim_path'],'sim'))
    else:
        # CurrentPath = sys.path[0]
        CurrentPath = './'
        SimPath = os.path.abspath(os.path.join(CurrentPath,'../sim'))
    print('''Do svn/git Operation''',flush = True)
    if GeneralDict['tools']=='svn':
        if GeneralDict['operation']=='co':
            if os.path.exists(RegressPath) is False:
                os.makedirs(RegressPath)
            os.chdir(RegressPath)
            url_list = ast.literal_eval(GeneralDict['url'])
            for url in url_list:
                os.chdir(RegressPath)
                CoSvnUrl(url)
            os.chdir(SimPath)
        elif GeneralDict['operation']=='update':
            os.chdir(SimPath)
            CallOperate("svn up",'Do svn Update Failed!!!')
    elif GeneralDict['tools']=='git':
        if GeneralDict['operation']=='clone':
            if os.path.exists(RegressPath) is False:
                os.makedirs(RegressPath)
            os.chdir(RegressPath)
            GitUrlListDict = ast.literal_eval(GeneralDict['url'])
            for url in GitUrlListDict.keys():
                print("git clone -b {} {}".format(GitUrlListDict[url],url))
                CallOperate("git clone -b {} {}".format(GitUrlListDict[url],url),'Do Git Clone Failed!!!')
        elif GeneralDict['operation']=='update':
            os.chdir(SimPath)
            CallOperate("git pull origin {}".format(GeneralDict['git_branch']),'Do Git Update Failed!!!')
    else:
        print("thie GeneralDict['tools']={} is illegal".format(GeneralDict['tools']))
    DoSourcePrj(SimPath)
    print('''Generate Daily Regress Task ''',flush = True)
    GenerateDailyRegress(GeneralDict,SimPath)
    return SimPath


def GenerateDailyRegress(GeneralDict,SimPath):
    if GeneralDict['set_daily_regress'] == 'enable':
        LocalPath = SimPath
        RegrShName = 'RegressCrontabScr'
        try:
            RegrIniFileName = sys.argv[1]
        except:
            RegrIniFileName = "regress.ini"
        RegrIniFile = IniPathMatch(RegrIniFileName)
        RegrIniFile = os.path.split(RegrIniFile)[1]
        RegrIniFile = RegrIniFile.split('.')[0]
        CshFile = GeneralDict['project_source_file']
        RegressHour   = int(GeneralDict['regress_date_hour'])
        RegressMinute = int(GeneralDict['regress_date_minute'])
        if RegressHour >= 24 or RegressHour < 0:
            RegressHour = 0
        if RegressMinute >= 60 or RegressMinute < 0:
            RegressMinute = 0
        RegressCron = CrontabMethod('daily_regress')
        RegressCron.GenLocalShScript(LocalPath,RegrIniFile,RegrShName,CshFile)
        RegressCron.SetCronTabByClock(RegressCron.GenRegressCmd(LocalPath,RegrShName),RegressHour,RegressMinute)
        RegressCron.PrintAllJob()


def DoPreRegress(SimPath,PreRegressDict):
    print('''pre_script''',flush = True)
    os.chdir(SimPath)
    if PreRegressDict['pre_script_enable']=='enable':
        PreScriptList = re.sub(" +"," ",PreRegressDict['pre_script']).split(';')
        for PreScriptCmd in PreScriptList:
            if len(PreScriptCmd) != 0:
                CallOperate(PreScriptCmd,'Do pre_script Failed')
    print('''pre_make''',flush = True)
    os.chdir(SimPath)
    if PreRegressDict['pre_make_enable']=='enable':
        PreMakeList = re.sub(" +"," ",PreRegressDict['pre_make']).split(';')
        for PreMakeCmd in PreMakeList:
            if len(PreMakeCmd) != 0:
                CallOperate(PreMakeCmd,'Do pre_make Failed')


def DoPostRegress(SimPath,PreRegressDict):
    print('''post_script''',flush = True)
    os.chdir(SimPath)
    if PreRegressDict['post_script_enable']=='enable':
        PostScriptList = re.sub(" +"," ",PreRegressDict['post_script']).split(';')
        for PostScriptCmd in PostScriptList:
            if len(PostScriptCmd) != 0:
                CallOperate(PostScriptCmd,'Do post_script Failed')
    print('''post_make''',flush = True)
    os.chdir(SimPath)
    if PreRegressDict['post_make_enable']=='enable':
        PostMakeList = re.sub(" +"," ",PreRegressDict['post_make']).split(';')
        for PostMakeCmd in PostMakeList:
            if len(PostMakeCmd) != 0:
                CallOperate(PostMakeCmd,'Do post_make Failed')


def GetCompileOpts(RegressDict):
    opts = " "
    opts += ' fcov={} '.format(RegressDict['fcov']) if 'fcov' in RegressDict.keys() else 'off'
    opts += ' ccov={} '.format(RegressDict['ccov']) if 'ccov' in RegressDict.keys() else 'off'
    opts += ' udf={}  '.format(RegressDict['udf'] ) if 'udf'  in RegressDict.keys() else ''
    opts += ' {} '.format(RegressDict['cmp_opts'] ) if 'cmp_opts' in RegressDict.keys() else ''
    return opts


def DoCompile(SimPath,ModeList,RegressDict):
    global RegressTime
    print( '''Do Compile''',flush = True)
    if RegressDict['docompile']=='enable':
        os.chdir(SimPath)
        for mode in ModeList:
            #MakeCmd = 'make clean mode={}'.format(mode)
            #CallOperate(MakeCmd,'Compile VCS with Command({}) wrong'.format(MakeCmd))
            MakeCmd = 'make compile mode={} {}'.format(mode,GetCompileOpts(RegressDict))
            CallOperate(MakeCmd,'Compile with Command({}) wrong'.format(MakeCmd))
    cmdFile = open("{}/cmd.file".format(SimPath),"w")
    cmdFile.write(RegressTime + '\n')
    cmdFile.write(str(ModeList) + '\n')
    cmdFile.write('{}'.format(RegressDict['ccov'] if 'ccov' in RegressDict.keys() else 'off') + '\n')
    cmdFile.write('{}'.format(RegressDict['fcov'] if 'fcov' in RegressDict.keys() else 'off') + '\n')
    cmdFile.write('{}'.format(RegressDict['fcov_key_word'] if 'fcov_key_word' in RegressDict.keys() else '_common_pkg') + '\n')
    cmdFile.write('{}'.format(RegressDict['dotransport'] if 'dotransport' in RegressDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(RegressDict['cmp_opts'] if 'cmp_opts' in RegressDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(RegressDict['dorotation'] if 'dorotation' in RegressDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(RegressDict['transport_while_regress'] if 'transport_while_regress' in RegressDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(RegressDict['del_pass_log'] if 'del_pass_log' in RegressDict.keys() else 'enable') + '\n')
    cmdFile.write('{}'.format(RegressDict['del_err_cov'] if 'del_err_cov' in RegressDict.keys() else 'enable') + '\n')
    cmdFile.close()


def GetTcList(RegressDict,):
    print('''Get TC List ''')
    TcList = []
    ModeList = []
    RandSeedList = []
    PreTcList = RegressDict['tc_list'].replace('[','').replace(']','').replace('\n','').replace('\t','').split(';')
    if PreTcList[-1]=='':
        PreTcList.pop(-1)
    if RegressDict['dorotation'] == 'enable':
        MaxRotationNum = -1
        TcDictList = []
        for tc in PreTcList:
            TcDict = ast.literal_eval(tc)
            #modeList
            if 'MODE' not in TcDict.keys():
                if 'base_fun' not in ModeList:
                    ModeList.append('base_fun')
            else:
                if TcDict['MODE'] not in ModeList:
                    ModeList.append(TcDict['MODE'])
            TcDict['RUN_TIMES'] = TcDict['RUN_TIMES'] if 'RUN_TIMES' in TcDict.keys() else 1
            TcDictList.append(TcDict)
            MaxRotationNum = max(TcDict['RUN_TIMES'],MaxRotationNum)
        for ronum in range(MaxRotationNum):
            for TcDict in TcDictList:
                RunNum = TcDict['RUN_TIMES']
                if RunNum != 0:
                    opts = ' '
                    opts += GetCompileOpts(RegressDict)
                    if 'SEED' in TcDict.keys():
                        if TcDict['SEED'] == 'rand':
                            SeedNum = random.randint(0,99999999)
                            while SeedNum in RandSeedList:
                                SeedNum = random.randint(0,99999999)
                            RandSeedList.append(SeedNum)
                            opts += ' seed={} '.format(SeedNum)
                        else:
                            opts += ' seed={} '.format(int(TcDict['SEED']))
                            RandSeedList.append(int(TcDict['SEED']))
                    if 'TC' not in TcDict.keys():
                        print('Error, Please Give Me the TC Name!!!')
                        sys.exit()
                    else:
                        opts += ' tc={} '.format(TcDict['TC'])
                    if 'MODE' not in TcDict.keys():
                        opts += ' mode=base_fun'
                    else:
                        opts += ' mode={}'.format(TcDict['MODE'])
                    opts += ' owner={} '.format(TcDict['TEST_OWNER']) if 'TEST_OWNER' in TcDict.keys() else ''
                    opts += ' {} '.format(TcDict['MAKE_OPTS']) if 'MAKE_OPTS' in TcDict.keys() else ''
                    TcList.append('make batch_run {}'.format(opts))
                    TcDict['RUN_TIMES'] -= 1
    else:
        for tc in PreTcList:
            TcDict = ast.literal_eval(tc)
            #modeList
            if 'MODE' not in TcDict.keys():
                if 'base_fun' not in ModeList:
                    ModeList.append('base_fun')
            else:
                if TcDict['MODE'] not in ModeList:
                    ModeList.append(TcDict['MODE'])
            #tcList
            RunNum = TcDict['RUN_TIMES'] if 'RUN_TIMES' in TcDict.keys() else 1
            for num in range(RunNum):
                opts = ' '
                opts += GetCompileOpts(RegressDict)
                if 'SEED' in TcDict.keys():
                    if TcDict['SEED'] == 'rand':
                        SeedNum = random.randint(0,99999999)
                        while SeedNum in RandSeedList:
                            SeedNum = random.randint(0,99999999)
                        RandSeedList.append(SeedNum)
                        opts += ' seed={} '.format(SeedNum)
                    else:
                        opts += ' seed={} '.format(int(TcDict['SEED']))
                        RandSeedList.append(int(TcDict['SEED']))
                if 'TC' not in TcDict.keys():
                    print('Error, Please Give Me the TC Name!!!')
                    sys.exit()
                else:
                    opts += ' tc={} '.format(TcDict['TC'])
                if 'MODE' not in TcDict.keys():
                    opts += ' mode=base_fun'
                else:
                    opts += ' mode={}'.format(TcDict['MODE'])
                opts += ' owner={} '.format(TcDict['TEST_OWNER']) if 'TEST_OWNER' in TcDict.keys() else ''
                opts += ' {} '.format(TcDict['MAKE_OPTS']) if 'MAKE_OPTS' in TcDict.keys() else ''
                TcList.append('make batch_run {}'.format(opts))
    return ModeList,TcList


def DelOneCov(Cmd,SimPath,SimTools):
    tcOptList = re.sub(' +',' ',Cmd).split(' ')
    TcName = ''
    ModeName = ''
    SeedNum = ''
    for Opt in tcOptList:
        if re.compile('tc=.*').match(Opt):
            TcName = Opt[3:]
        if re.compile('seed=.*').match(Opt):
            SeedNum = Opt[5:]
        if re.compile('mode=.*').match(Opt):
            ModeName = Opt[5:]
    if SimTools == 'vcs':
        CovPath = "{}/{}/cov/simv.vdb/vcs_cov/simv_rtl.cm.vdb/snps/coverage/db/testdata/{}_{}".format(SimPath,ModeName,TcName,SeedNum)
    elif SimTools == 'xrun':
        CovPath = "{}/{}/cov/top_tb/{}_{}".format(SimPath,ModeName,TcName,SeedNum)
    if os.path.exists(CovPath) is True:
        os.system("rm -rf {}".format(CovPath))
    return


def DoOneSimv(MakeCmd,InfoDict,GeneralDict,TcIdx,TotalTcNum,SimPath,RemainCmdList):
    # CallOperate(MakeCmd,'Runnint Tc with Command({}) wrong'.format(MakeCmd))----TODO----why the udr work wrong(udr=+frame_num=1 but the tc had detect the frame_num is 0)
    # os.system(MakeCmd)
    print('''({})Running Tc({}/{}) with Command({})......'''.format(time.strftime("%Y%m%d_%H%M%S",time.localtime()),TcIdx,TotalTcNum,MakeCmd),flush = True)
    os.system(MakeCmd + ">/dev/null 2>&1")
    fail_recmp_flag = GeneralDict['fail_recmp']
    fail_rerun_flag = GeneralDict['fail_rerun']
    tcLogName = GenLogName(MakeCmd)
    logResult,CpuTimeInfo,MemoryInfo,NotPassLine = DoOneLogScan(tcLogName,InfoDict['KewWordDict'],[InfoDict['StartChkFatalWord'],InfoDict['StartChkErrorWord'],InfoDict['StartChkWarningWord']],InfoDict['SimTools'])
    # delete error coverage
    if(InfoDict['DelErrCov'] == 'enable') and (logResult == 'Fatal' or logResult == 'Error') and (InfoDict['CcovEn'] == 'on' or InfoDict['FcovEn'] == 'on'):
        DelOneCov(MakeCmd,SimPath,InfoDict['SimTools'])
    global rerunTimes
    if (logResult == 'Fatal' or logResult == 'Error') and (rerunTimes < 1):
        fail_recmp_opt = GeneralDict['recmp_opts']
        fail_rerun_opt = GeneralDict['rerun_opts']
        if fail_rerun_flag == 'enable':
            print("fail rerun or recompile " ,flush = True)
            if fail_recmp_flag == 'disable':
                RerunCmd = MakeCmd+ ' ' + fail_rerun_opt
            else:
                # rerun and recompile
                MakeCmd = re.sub('fcov={}'.format(InfoDict['FcovEn']),'',MakeCmd)
                MakeCmd = re.sub('ccov={}'.format(InfoDict['CcovEn']),'',MakeCmd)
                MakeCmd = re.sub('{}'.format(InfoDict['CmpOptions']),'',MakeCmd)
                MakeCmd = re.sub(r'udf=".*" ','',MakeCmd)
                RerunCmd = MakeCmd+ ' ' + fail_recmp_opt + ' ' + fail_rerun_opt + ' simv_by_tc=on'
                RerunCmd = RerunCmd.replace('batch_run','run',1)
            print("rerun cmd: ",RerunCmd)
            # CallOperate(RerunCmd, 'Rerun Tc with Command({}) wrong'.format(RerunCmd))
            os.system(RerunCmd)
            rerunTimes += 1
    print('''({})Run Tc({}/{}) with Command({}) Done!'''.format(time.strftime("%Y%m%d_%H%M%S",time.localtime()),TcIdx,TotalTcNum,MakeCmd),flush = True)
    if TcIdx + 1 == TotalTcNum:
        GenRemainIni([],SimPath)
    else:
        GenRemainIni(RemainCmdList[TcIdx:],SimPath)


def GenLogName(tc):
    tcOptList = re.sub(' +',' ',tc).split(' ')
    LogName = ''
    TcName = ''
    SeedNum = ''
    ModeName = ''
    NoteName = ''
    TimingOpt = 'rtl'
    NoteOpt = ''
    for Opt in tcOptList:
        if re.compile('tc=.*').match(Opt):
            TcName = Opt[3:]
        if re.compile('seed=.*').match(Opt):
            SeedNum = Opt[5:]
        if re.compile('mode=.*').match(Opt):
            ModeName = Opt[5:]
        if re.compile('note=.*').match(Opt):
            NoteName = Opt[5:]
        if re.compile('timing=.*').match(Opt):
            TimingOpt = Opt[7:]
        if re.compile('note=.*').match(Opt):
            NoteOpt = "_" + Opt[5:]
    if NoteOpt != '':
        LogName = ModeName + "/log/" + TcName + "_" + SeedNum + "_" + TimingOpt + NoteOpt + ".log"
    else:
        LogName = ModeName + "/log/" + TcName + "_" + SeedNum + "_" + TimingOpt + NoteOpt + ".log"
    # print('{}'.format(LogName))
    return LogName


def GetOneCmdString(InfoDict,cmd):
    tmpCmd = cmd.split(' ')
    tmpMakeOpts = ""
    for ele in tmpCmd:
        if 'seed=' in ele:
            tmpEle = ele.split('=')
            tmpSeed = tmpEle[1]
        elif 'tc=' in ele:
            tmpEle = ele.split('=')
            tmpTc = tmpEle[1]
        elif 'mode=' in ele:
            tmpEle = ele.split('=')
            tmpMode = tmpEle[1]
        elif 'owner=' in ele:
            tmpEle = ele.split('=')
            tmpOwner = tmpEle[1]
        elif 'make' in ele:
            continue
            #do nothing
        elif 'run' in ele:
            continue
            #do nothing
        elif 'fcov' in ele:
            continue
            #do nothing
        elif 'ccov' in ele:
            continue
            #do nothing
        elif 'udf' in ele:
            continue
            #do nothing
        else:
            # make opts
            if ele not in InfoDict['CmpOptions']:
                tmpMakeOpts += " "
                tmpMakeOpts += ele
    tmpTcList = ""
    tmpTcList += '{\'RUN_TIMES\':1,'
    tmpTcList += '\'TC\':\'{}\','.format(tmpTc)
    tmpTcList += '\'MODE\':\'{}\','.format(tmpMode)
    tmpTcList += '\'SEED\':\'{}\','.format(tmpSeed)
    tmpTcList += '\'TEST_OWNER\':\'{}\','.format(tmpOwner)
    tmpTcList += '\'MAKE_OPTS\':\'{}\''.format(tmpMakeOpts)
    tmpTcList += ' };\n'
    return tmpTcList


def Cmd2Tclist(InfoDict,CmdList):
    TcList = []
    for Cmd in CmdList:
        TcList.append(GetOneCmdString(InfoDict,Cmd))
    return TcList


def GenRemainIni(CmdList,SimPath):
    tmpInfoDict = GetRegressCfgInfo(SimPath)
    tmpTcList = Cmd2Tclist(tmpInfoDict,CmdList)
    GenNewIni(tmpInfoDict,tmpTcList,SimPath,'../regress/RemainTc.ini','disable')


def DoAllSimv(SimPath,TcList,GeneralDict):
    global RegressTim
    print('''Do All Simv''',flush = True)
    cmdFile = open("{}/cmd.file".format(SimPath),"a")
    cmdFile.write('{}'.format(GeneralDict['send_to_workwx'] if 'send_to_workwx' in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['send_to_email' ] if 'send_to_email'  in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['receiver_list' ] if 'receiver_list'  in GeneralDict.keys() else '[]') + '\n')
    cmdFile.write('{}'.format(GeneralDict['tools'] if 'tools' in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['ignore_warning_en'] if 'ignore_warning_en' in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['fatal_key_words'] if 'fatal_key_words' in GeneralDict.keys() else '[]') + '\n')
    cmdFile.write('{}'.format(GeneralDict['error_key_words'] if 'error_key_words' in GeneralDict.keys() else '[]') + '\n')
    cmdFile.write('{}'.format(GeneralDict['warning_key_words'] if 'warning_key_words' in GeneralDict.keys() else '[]') + '\n')
    cmdFile.write('{}'.format(GeneralDict['ignore_fatal_key_words'] if 'ignore_fatal_key_words' in GeneralDict.keys() else '{}') + '\n')
    cmdFile.write('{}'.format(GeneralDict['ignore_error_key_words'] if 'ignore_error_key_words' in GeneralDict.keys() else '{}') + '\n')
    cmdFile.write('{}'.format(GeneralDict['ignore_warning_key_words'] if 'ignore_warning_key_words' in GeneralDict.keys() else '{}') + '\n')
    cmdFile.write('{}'.format(GeneralDict['total_ignore_log_list'] if 'total_ignore_log_list' in GeneralDict.keys() else '[]') + '\n')
    cmdFile.write('{}'.format(GeneralDict['start_chk_fatal_key_word'  ] if 'start_chk_fatal_key_word'   in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['start_chk_error_key_word'  ] if 'start_chk_error_key_word'   in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['start_chk_warning_key_word'] if 'start_chk_warning_key_word' in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['module_name'] if 'module_name' in GeneralDict.keys() else 'dut') + '\n')
    cmdFile.write('{}'.format(GeneralDict['excel_name'] if 'excel_name' in GeneralDict.keys() else 'dut_excel') + '\n')
    cmdFile.write('{}'.format(GeneralDict['excel_path'] if 'excel_path' in GeneralDict.keys() else './') + '\n')
    cmdFile.write('{}'.format(GeneralDict['sim_tools'] if 'sim_tools' in GeneralDict.keys() else 'vcs') + '\n')
    cmdFile.write('{}'.format(GeneralDict['tapd_enable'] if 'tapd_enable' in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['tapd_module_name'] if 'tapd_module_name' in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['anti_annotation_enable'] if 'anti_annotation_enable' in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['testpoint_file_name'] if 'testpoint_file_name' in GeneralDict.keys() else '') + '\n')
    cmdFile.write('{}'.format(GeneralDict['set_daily_regress'] if 'set_daily_regress' in GeneralDict.keys() else 'disable') + '\n')
    cmdFile.write('{}'.format(GeneralDict['regress_date_hour'] if 'regress_date_hour' in GeneralDict.keys() else '0') + '\n')
    cmdFile.write('{}'.format(GeneralDict['regress_date_minute'] if 'regress_date_minute' in GeneralDict.keys() else '0') + '\n')
    cmdFile.write('{}'.format(GeneralDict['project_source_file'] if 'project_source_file' in GeneralDict.keys() else '') + '\n')
    for tc in TcList:
        tcLogName = GenLogName(tc)
        cmdFile.write(tcLogName + " " + tc + '\n')
    cmdFile.close()
    GenRemainIni(TcList,SimPath)
    if os.path.exists("{}/cmd_dir".format(SimPath)) is False:
        os.makedirs("{}/cmd_dir".format(SimPath))
    os.system("cp {}/cmd.file {}/cmd_dir/cmd_{}.file".format(SimPath,SimPath,RegressTime))
    os.chdir(SimPath)
    InfoDict = GetRegressCfgInfo(SimPath)
    p = multiprocessing.Pool(int(GeneralDict['paral_run_num']))
    total_tc_num = len(TcList)
    tc_idx = 0
    for tc in TcList:
        p.apply_async(DoOneSimv,args=(tc,InfoDict,GeneralDict,tc_idx,total_tc_num,SimPath,TcList))
        tc_idx = tc_idx + 1
    p.close()
    p.join()
    os.remove(os.path.join(SimPath,"../regress/RemainTc.ini"))
    print("======================================================",flush = True)
    print("All TC Regression Done.",flush = True)
    print("======================================================",flush = True)


def DoRegress(SimPath,GeneralDict,PreRegressDict,RegressDict):
    tmp_ModeList,tmp_TcList = GetTcList(RegressDict)
    DoCompile(SimPath,tmp_ModeList,RegressDict)
    DoAllSimv(SimPath,tmp_TcList,GeneralDict)


def KeyWordListAppend(OrgnList, AppendList):
    for keyword in AppendList:
        if keyword not in OrgnList:
            OrgnList.append(keyword)


def KeyWordListIgnore(OrgnList, IgnoreList):
    for ignoreword in IgnoreList:
        for keyword in OrgnList:
            if ignoreword==keyword:
                OrgnList.remove(ignoreword)
                break


def GetRegressCfgInfo(SimPath):
    InfoDict = {}
    os.chdir(SimPath)
    with open("./cmd.file", "r") as cmdFile:
        tmpCmds = cmdFile.readlines()
        cmdFile.close()
    cmdsList = []
    for cmd in tmpCmds:
        cmd = cmd.replace('\n','')
        cmdsList.append(cmd)
    InfoDict['StartRegressTime'] = cmdsList.pop(0)
    InfoDict['ModeList'] = eval(cmdsList.pop(0))
    InfoDict['CcovEn'] = cmdsList.pop(0)
    InfoDict['FcovEn'] = cmdsList.pop(0)
    InfoDict['FcovKeyWord'] = cmdsList.pop(0)
    InfoDict['DoTransPort'] = cmdsList.pop(0)
    InfoDict['CmpOptions'] = cmdsList.pop(0)
    InfoDict['DoRotation'] = cmdsList.pop(0)
    InfoDict['TransportWhileRegress'] = cmdsList.pop(0)
    InfoDict['DelPassLog'] = cmdsList.pop(0)
    InfoDict['DelErrCov'] = cmdsList.pop(0)
    InfoDict['Send2Workwx'] = cmdsList.pop(0)
    InfoDict['Send2Email'] = cmdsList.pop(0)
    InfoDict['ReceiverList'] = eval(cmdsList.pop(0))
    InfoDict['ToolInfo'] = cmdsList.pop(0)
    is_ignore_warning_en = cmdsList.pop(0)
    FatalKeyWordList = ["UVM_FATAL"]
    ErrorKeyWordList = ["UVM_ERROR","fail","failed","Error"]
    WarningKeyWordList = ["UVM_WARNING","warning"]
    tmpFatalKeyWordList = eval(cmdsList.pop(0))
    tmpErrorKeyWordList = eval(cmdsList.pop(0))
    tmpWarningKeyWordList = eval(cmdsList.pop(0))
    IgnoreFatalKeyWordDict = eval(cmdsList.pop(0))
    IgnoreErrorKeyWordDict = eval(cmdsList.pop(0))
    IgnoreWarningKeyWordDict = eval(cmdsList.pop(0))
    KeyWordListAppend(FatalKeyWordList,tmpFatalKeyWordList)
    KeyWordListAppend(ErrorKeyWordList,tmpErrorKeyWordList)
    KeyWordListAppend(WarningKeyWordList,tmpWarningKeyWordList)
    if is_ignore_warning_en=='enable':
        WarningKeyWordList = []
    KewWordDictValue = {'FatalKeyWordList':FatalKeyWordList,'ErrorKeyWordList':ErrorKeyWordList,'WarningKeyWordList':WarningKeyWordList,'IgnoreFatalDict':IgnoreFatalKeyWordDict,'IgnoreErrorDict':IgnoreErrorKeyWordDict,'IgnoreWarningDict':IgnoreWarningKeyWordDict}
    InfoDict['KewWordDict'] = KewWordDictValue
    TotalIgnoreLogList = ['vcs_compile_rtl.log','git.log','.*.log.swp','.*.log.swo']
    tmpTotalIgnoreLogList = eval(cmdsList.pop(0))
    KeyWordListAppend(TotalIgnoreLogList,tmpTotalIgnoreLogList)
    InfoDict['TotalIgnoreLogList'] = TotalIgnoreLogList
    InfoDict['StartChkFatalWord']   = cmdsList.pop(0)
    InfoDict['StartChkErrorWord']   = cmdsList.pop(0)
    InfoDict['StartChkWarningWord'] = cmdsList.pop(0)
    InfoDict['CovInstanceName'] = cmdsList.pop(0)
    InfoDict['ExcelName'] = cmdsList.pop(0)
    InfoDict['ExcelPath'] = cmdsList.pop(0)
    InfoDict['SimTools'] = cmdsList.pop(0)
    InfoDict['TapdEnable'] = cmdsList.pop(0)
    InfoDict['TapdModuleName'] = cmdsList.pop(0)
    InfoDict['AntiAnnotationEnable'] = cmdsList.pop(0)
    InfoDict['TestpointFileName'] = cmdsList.pop(0)
    InfoDict['SetDailyRegress'] = cmdsList.pop(0)
    InfoDict['RegressDateHour'] = cmdsList.pop(0)
    InfoDict['RegressDateMinute'] = cmdsList.pop(0)
    InfoDict['ProjectSourceFile'] = cmdsList.pop(0)
    ExecuteLogList = []
    ExecuteCmdList = []
    for cmd in cmdsList:
        tmpCmd = cmd.split(' ',1)
        ExecuteLogList.append(tmpCmd[0])
        ExecuteCmdList.append(re.sub(" +"," ",tmpCmd[1]))
    InfoDict['ExecuteLogList'] = ExecuteLogList
    InfoDict['ExecuteCmdList'] = ExecuteCmdList
    return InfoDict


def GetModeLogList(TotalIgnoreLogList,ModePath,Mode):
    ModeLogList = []
    ModePath += '/log'
    tmp_list = os.listdir(ModePath)
    for log in tmp_list:
        if log not in TotalIgnoreLogList:
            tmp_log = Mode + '/log/' + log
            ModeLogList.append(tmp_log)
    return ModeLogList


def GetTotalLogList(SimPath,TotalIgnoreLogList,ModeList):
    TotalLogList = []
    for Mode in ModeList:
        ModePath = SimPath+'/'+Mode
        if not os.path.exists(ModePath):
            print("There is no ModePath named: ",ModePath)
        else:
            tmp_list = GetModeLogList(TotalIgnoreLogList,ModePath,Mode)
            TotalLogList.extend(tmp_list)
    return TotalLogList


def DoLineSearchWord(line,KeyWordList):
    Result = False
    for word in KeyWordList:
        # if re.compile(word,re.I).search(line):
        if re.compile(word).search(line):
            Result = True
            break
    return Result


def DoLineSearchWordDict(line,KeyWordDict):
    Result = False
    if type(KeyWordDict) == type([]):
        for word in KeyWordDict:
            if re.compile(word).search(line):
                Result = True
                break
    elif type(KeyWordDict) == type({}):
        for word in KeyWordDict.keys():
            #if re.compile(word,re.I).search(line):
            if re.compile(word).search(line): 
                for value in KeyWordDict[word]:
                    if re.compile(value).search(line) :
                        Result = True
                        break
    return Result


def DoOneLogScan(log,KewWordDict,StartChkWordList,SimTools):
    FatalKeyWordList = KewWordDict["FatalKeyWordList"]
    ErrorKeyWordList = KewWordDict["ErrorKeyWordList"]
    WarningKeyWordList = KewWordDict["WarningKeyWordList"]
    IgnoreFatalDict   = KewWordDict["IgnoreFatalDict"]
    IgnoreErrorDict   = KewWordDict["IgnoreErrorDict"]
    IgnoreWarningDict = KewWordDict["IgnoreWarningDict"]
    StartChkFatalWord = StartChkWordList[0]
    StartChkErrorWord = StartChkWordList[1]
    StartChkWarningWord = StartChkWordList[2]
    EndKeyWordList = ["TEST CASE PASSED","TEST CASE FAILED","UVM Report catcher Summary","UVM Report Summary"]
    CpuTimeKeyWordList = ["CPU Time","xrun: Time"]
    CpuMemKeyWordList = ["Data structure size","xmsim: Memory Usage - Final:"]
    logResult = 'NotStart' #['NotStart','Running','Fatal','Error','Warning','Pass']
    CpuTimeInfo = '0 seconds'
    MemoryInfo = '0 Mb'
    NotPassLine = ''
    ChkFatalFlag = False
    ChkErrorFlag = False
    ChkWarningFlag = False
    if os.path.exists(log)==False:
        return logResult,CpuTimeInfo,MemoryInfo,NotPassLine
    logResult = 'Running'
    with open(log, "r") as f:
        lines = f.readlines()
        IsEndOfLog = False
        for line in lines:
            if IsEndOfLog==False:
                '''search end of log and break the log analysis if match'''
                if DoLineSearchWord(line,EndKeyWordList)==True:
                    IsEndOfLog = True
                    if logResult=='Running':
                        logResult = 'Pass'
            if IsEndOfLog==False:
                if DoLineSearchWord(line,[StartChkFatalWord]) == True or StartChkFatalWord == '':
                    ChkFatalFlag = True
                if DoLineSearchWord(line,[StartChkErrorWord]) == True or StartChkErrorWord == '':
                    ChkErrorFlag = True
                if DoLineSearchWord(line,[StartChkWarningWord]) == True or StartChkFatalWord == '':
                    ChkWarningFlag = True
                if logResult=='Running':
                    '''search warning and record the warning quantity if match'''
                    if DoLineSearchWord(line,WarningKeyWordList)==True and DoLineSearchWordDict(line,IgnoreWarningDict)==False and ChkWarningFlag == True:
                        logResult = 'Warning'
                        NotPassLine = str(line).replace('\n','')
                if logResult in ['Running','Warning']:
                    '''search error and record the error quantity if match'''
                    if DoLineSearchWord(line,ErrorKeyWordList)==True and DoLineSearchWordDict(line,IgnoreErrorDict)==False and ChkErrorFlag == True:
                        logResult = 'Error'
                        NotPassLine = str(line).replace('\n','')
                if logResult in ['Running','Warning','Error']:
                    '''search fatal and break the log analysis if match'''
                    if DoLineSearchWord(line,FatalKeyWordList)==True and DoLineSearchWordDict(line,IgnoreFatalDict)==False and ChkFatalFlag == True:
                        logResult = 'Fatal'
                        NotPassLine = str(line).replace('\n','')
            '''search cpu time info'''
            if DoLineSearchWord(line,CpuTimeKeyWordList)==True:
                if SimTools == 'vcs':
                    tmpline = line.replace(';',' ').split()
                    for strele in tmpline:
                        if strele == 'seconds':
                            CpuTimeInfo = laststr + ' seconds'
                        else:
                            laststr = strele
                elif SimTools == 'xrun':
                    tmpline = line.replace('\n','').split(' ')
                    if tmpline[-1] != '':
                        CpuTimeInfo = tmpline[-1].replace('s','seconds')
                else:
                    print('Your sim tools setting ({}) is error, so CPU Time will not print!'.format(SimTools))
            '''search memory info'''
            if DoLineSearchWord(line,CpuMemKeyWordList)==True:
                if SimTools == 'vcs':
                    tmpline = line.replace(';',' ').replace('\n','').split()
                    MemoryInfo = tmpline[-1] if 'Mb' in tmpline[-1] else '0 Mb'
                elif SimTools == 'xrun':
                    MemKeyWord = 'Peak:'
                    tmpline = line.replace('\n','').replace(',','').split(' ')
                    laststr = ''
                    for strele in tmpline:
                        if strele == MemKeyWord:
                            MemoryInfo = laststr
                        else :
                            laststr = strele
                else:
                    print('Your sim tools setting ({}) is error, so Mem Cost will not print!'.format(SimTools))
        f.close()
    return logResult,CpuTimeInfo,MemoryInfo,NotPassLine


def DoLogScan(LogList,KewWordDict,DelPassLog,StartChkWordList,SimTools):
    FatalNum,ErrorNum,WarningNum,RunningNum,NotStartNum,PassNum = 0,0,0,0,0,0
    TotalCpuTime = 0
    MaxCpuTime,MaxMemory = 0,0
    MaxCpuTimeInfo,MaxMemoryInfo = ['NULL','NULL','0 seconds','0 Mb',''],['NULL','NULL','0 seconds','0 Mb','']
    LogInfoList = []
    for log in LogList:
        logResult,CpuTimeInfo,MemoryInfo,NotPassLine = DoOneLogScan(log,KewWordDict,StartChkWordList,SimTools)
        if logResult=='Fatal':
            FatalNum = FatalNum + 1
        elif logResult=='Error':
            ErrorNum = ErrorNum + 1
        elif logResult=='Warning':
            WarningNum = WarningNum + 1
        elif logResult=='Running':
            RunningNum = RunningNum + 1
        elif logResult=='NotStart':
            NotStartNum = NotStartNum + 1
        else:
            PassNum = PassNum + 1
            # delete pass log
            if DelPassLog == "enable":
                os.remove(log)
        tmpList = [os.path.join('..',str(log)),logResult,CpuTimeInfo,MemoryInfo,NotPassLine]
        LogInfoList.append(tmpList)
        tmpCpuTimeValue = CpuTimeInfo.split("seconds")
        CpuTimeValue = float(tmpCpuTimeValue[0].replace(" ",""))
        TotalCpuTime = TotalCpuTime + CpuTimeValue
        if CpuTimeValue > MaxCpuTime:
            MaxCpuTime = CpuTimeValue
            MaxCpuTimeInfo = tmpList
        tmpMemoryValue = MemoryInfo.split("M")
        MemoryValue = float(tmpMemoryValue[0].replace(" ",""))
        if MemoryValue > MaxMemory:
            MaxMemory = MemoryValue
            MaxMemoryInfo = tmpList
    LogResultDict = {}
    LogResultDict['FatalNum'] = FatalNum
    LogResultDict['ErrorNum'] = ErrorNum
    LogResultDict['WarningNum'] = WarningNum
    LogResultDict['RunningNum'] = RunningNum
    LogResultDict['NotStartNum'] = NotStartNum
    LogResultDict['PassNum'] = PassNum
    LogResultDict['TotalCpuTime'] = "{0:.2f}".format(TotalCpuTime)
    LogResultDict['MaxCpuTimeInfo'] = MaxCpuTimeInfo
    LogResultDict['MaxMemoryInfo'] = MaxMemoryInfo
    LogResultDict['LogInfoList'] = LogInfoList
    return LogResultDict


def VCSCcovResult(SimPath,ModuleName):
    CcovPath = os.path.abspath(os.path.join(SimPath,'urgReport/hierarchy.txt'))
    CcovResult = []
    if os.path.exists(CcovPath):
        ccov_txt = open(CcovPath)
        ccov_lines = ccov_txt.readlines()
        if(ModuleName==''):
            CcovResult = ['--','--','--','--','--']
        else:
            last_item = ''
            tmpCovDic = {}
            for i in ccov_lines:
                if ModuleName in i:
                    label = last_item.split()
                    items=i.split()
                    for i in range(len(label)):
                        try:
                            tmpCovDic[label[i]] = format(float(items[i])/100,'.2%')
                        except:
                            tmpCovDic[label[i]] = items[i]
                    break;
                elif 'SCORE' in i or 'LINE' in i or 'COND' in i or 'TOGGLE' in i or 'FSM' in i:
                    last_item = i
            if 'SCORE' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['SCORE'])
            else:
                CcovResult.append('--')
            if 'LINE' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['LINE'])
            else:
                CcovResult.append('--')
            if 'COND' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['COND'])
            else:
                CcovResult.append('--')
            if 'TOGGLE' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['TOGGLE'])
            else:
                CcovResult.append('--')
            if 'FSM' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['FSM'])
            else:
                CcovResult.append('--')
    else:
        CcovResult = ['--','--','--','--','--']
        print('NO code coverage file generated')
    return CcovResult


def VCSFcovResult(SimPath):
    FcovPath = os.path.abspath(os.path.join(SimPath,'urgReport/groups.txt'))
    if os.path.exists(FcovPath):
        line_number = 0
        total_fcov_score_line = -1
        fcov_txt = open(FcovPath)
        fcov_lines = fcov_txt.readlines()
        for i in fcov_lines:
            line_number += 1
            if 'Total Groups Coverage Summary'  in i:
                total_fcov_score_line = line_number+2
            if line_number == total_fcov_score_line:
                item = i.split()
                try:
                    FcovResult = format(float(item[0])/100,'.2%')
                except:
                    FcovResult = item[0]
    else:
        FcovResult = '--'
    return FcovResult


def XRunCcovResult(SimPath,ModuleName,ModeName):
    CcovPath = os.path.abspath(os.path.join(SimPath,'{}_cov_report.txt'.format(ModeName)))
    if os.path.exists(CcovPath):
        ccov_txt = open(CcovPath)
        ccov_lines = ccov_txt.readlines()
        if(ModuleName==''):
            CcovResult = ['--','--','--','--','--']
        else:
            last_item = ''
            tmpCovDic = {}
            CcovResult = []
            for i in ccov_lines:
                if ModuleName in i:
                    label =['Overall', 'Block', 'Expression', 'Toggle', 'Fsm' ]
                    CcovResult = []
                    items = []
                    for ele in i.split():
                        if '%' in ele:
                            items.append(ele.replace('%',''))
                        if 'n/a' in ele:
                            items.append('--')
                    for i in range(len(label)):
                        try:
                            tmpCovDic[label[i]] = format(float(items[i])/100,'.2%')
                        except:
                            tmpCovDic[label[i]] = items[i]
                    break;
                elif 'Overall' in i or 'Block' in i or 'Expression' in i or 'Toggle' in i or 'Fsm' in i:
                    last_item = i
            if 'Overall' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['Overall'])
            else:
                CcovResult.append('--')
            if 'Block' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['Block'])
            else:
                CcovResult.append('--')
            if 'Expression' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['Expression'])
            else:
                CcovResult.append('--')
            if 'Toggle' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['Toggle'])
            else:
                CcovResult.append('--')
            if 'Fsm' in tmpCovDic.keys():
                CcovResult.append(tmpCovDic['Fsm'])
            else:
                CcovResult.append('--')
    else:
        CcovResult = ['--','--','--','--','--']
    return CcovResult


def XRunFcovResult(SimPath,ModeName,FcovKeyWord):
    FcovResult = '--'
    FcovPath = os.path.abspath(os.path.join(SimPath,'{}_cov_report.txt'.format(ModeName)))
    FcovModuleName = '_common_pkg' if FcovKeyWord == '' else FcovKeyWord
    if os.path.exists(FcovPath):
        fcov_txt = open(FcovPath)
        fcov_lines = fcov_txt.readlines()
        for line in fcov_lines:
            if FcovModuleName in line:
                tmplist = line.replace('\n','').replace('%','').split()
                try:
                    FcovResult = format(float(tmplist[-2])/100,'.2%')
                except:
                    FcovResult = '--'
                break
    else:
        FcovResult = '--'
    return FcovResult


def GetCovResult(SimPath,ModeList,CcovEn,FcovEn,CovInstanceName,SimTools,InfoDict):
    i_items = []
    CovResultDict = {}
    ModuleName = CovInstanceName
    '''Try Get ModuleName'''
    if CcovEn == 'on' or FcovEn == 'on':
        '''read coverage result in txt'''
        os.chdir(SimPath)
        for Mode in ModeList:
            CovResult = []
            get_cov_result = True
            try:
                CallOperate('make cov_txt mode={}'.format(Mode),'error make cov_txt')
            except:
                get_cov_result = False
            if get_cov_result==False:
                CovResult = ['--','--','--','--','--','--']
                CovResultDict[Mode] = CovResult
                continue
            if SimTools == 'vcs':
                CovResult = VCSCcovResult(SimPath,ModuleName)
                FcovResult = VCSFcovResult(SimPath)
            elif SimTools == 'xrun':
                CovResult = XRunCcovResult(SimPath,ModuleName,Mode)
                FcovResult = XRunFcovResult(SimPath,Mode,InfoDict['FcovKeyWord'])
            else:
                CovResult = ['--','--','--','--','--']
                FcovResult = '--'
            CovResult.append(FcovResult)
            CovResultDict[Mode] = CovResult
    else:
        for Mode in ModeList:
            CovResult = ['--','--','--','--','--','--']
            CovResultDict[Mode] = CovResult
    return CovResultDict


def GetProjectVersion(SimPath,ToolInfo):
    if ToolInfo=='git':
        try:
            # Python3
            GitBranch = subprocess.check_output("git symbolic-ref --short HEAD".split(' ')).decode().replace('\n','')
            GitCommitId = subprocess.check_output("git rev-parse HEAD".split(' ')).decode().replace('\n','')
            GitStatus = subprocess.check_output("git status -uno".split(' ')).decode()
        except TypeError:
            # Python2
            GitBranch = subprocess.check_output("git symbolic-ref --short HEAD".split(' ')).replace('\n','')
            GitCommitId = subprocess.check_output("git rev-parse HEAD".split(' ')).replace('\n','')
            GitStatus = subprocess.check_output("git status -uno".split(' '))
        #VersionInfo = ">"*4 + " branch:{} commit_id:{} ".format(GitBranch,GitCommitId).ljust(WidthOfLogName+WidthOfResult+1-4)
        VersionInfo = "git version info>>>>\n{} ".format(GitStatus)
        VersionInfo = VersionInfo[:-1]
        VersionInfo += "# commid_id: {}".format(GitCommitId)
    elif ToolInfo=='svn':
         os.chdir('../')
         svn_t = os.popen('svn info')
         svnvar = svn_t.read()
         svnList = svnvar.splitlines()
         SvnVersion = ""
         for svnstr in svnList:
             if svnstr.find('URL:') >= 0:
                 SvnVersion += "{}\n".format(svnstr) 
             elif svnstr.find('Repository Root:') >= 0:
                 SvnVersion += "{}\n".format(svnstr) 
             elif svnstr.find('Revision:') >= 0:
                 SvnVersion += "{}\n".format(svnstr)
         #VersionInfo = "SVN Version is:\n{} ".format(SvnVersion)
         VersionInfo = SvnVersion
         os.chdir(SimPath)
    else :
        VersionInfo = ""
    return VersionInfo


def GenRegressResultLog(SimPath,IsScanTotalEn,TotalNum,ExecuteNum,InfoDict,LogResultDict,CovResultDict,ProjectVersion,CmdList):
    EndRegressTime = time.strftime("%Y%m%d_%H%M%S",time.localtime())
    if IsScanTotalEn=='enable':
        log_name = "regress_result_total.log"
    else:
        log_name = "regress_result_{}.log".format(InfoDict['StartRegressTime'])
    WidthOfLogName = 85
    WidthOfResult = 16
    widthOfInfo = 20
    SplitFlag0 = "="*(WidthOfLogName+WidthOfResult+widthOfInfo*2+4) + "\n"
    SplitFlag1 = "-"*(WidthOfLogName+WidthOfResult+widthOfInfo*2+4) + "\n"
    Title0 = "REGRESS STATISTIC".center(WidthOfLogName+WidthOfResult+1) + "\n"
    Title1 = "TC LOG".center(WidthOfLogName) + "|" + "RESULT".center(WidthOfResult) + "|" + "CPU time info".center(widthOfInfo) + "|" + "Memory info".center(widthOfInfo) + "\n"
    Title2 = "COMMOND LIST".center(WidthOfLogName+WidthOfResult+1) + "\n"
    '''==================================================================VersionInfo================================================='''
    VersionInfo = ProjectVersion + '\n'
    '''==================================================================TimeInfo================================================='''
    if IsScanTotalEn=='enable':
        TimeInfo = "StartRegressTime : NULL\nEndRegressTime   : NULL\n"
    else:
        TimeInfo = "StartRegressTime : {}\nEndRegressTime   : {}\n".format(InfoDict['StartRegressTime'],EndRegressTime)
    TimeInfo = TimeInfo + "TOTAL CPUTIME : " + "{}".format(LogResultDict['TotalCpuTime']) + " seconds" + "\n"
    TimeInfo = TimeInfo + "MAX CPUTIME   : " + "{}".format(LogResultDict['MaxCpuTimeInfo'][0]).ljust(WidthOfLogName-16) + ":" + " {}".format(LogResultDict['MaxCpuTimeInfo'][2]).center(WidthOfResult) + "\n"
    TimeInfo = TimeInfo + "MAX MEMORY    : " + "{}".format(LogResultDict['MaxMemoryInfo'][0]).ljust(WidthOfLogName-16) + ":" + " {}".format(LogResultDict['MaxMemoryInfo'][3]).center(WidthOfResult) + "\n"
    '''==================================================================StatisInfo================================================='''
    StatisInfoList = []
    StatisInfoList.append("   TOTAL".ljust(WidthOfLogName) + ':' + "{}".format(TotalNum).center(WidthOfResult))
    StatisInfoList.append("   EXECUTED".ljust(WidthOfLogName) + ':' + "{}".format(ExecuteNum).center(WidthOfResult))
    StatisInfoList.append("   FATAl".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['FatalNum']).center(WidthOfResult))
    StatisInfoList.append("   ERROR".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['ErrorNum']).center(WidthOfResult))
    StatisInfoList.append("   WARNING".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['WarningNum']).center(WidthOfResult))
    StatisInfoList.append("   RUNNING".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['RunningNum']).center(WidthOfResult))
    StatisInfoList.append("   NOT START".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['NotStartNum']).center(WidthOfResult))
    StatisInfoList.append("   PASS".ljust(WidthOfLogName) + ':' + "{}".format(LogResultDict['PassNum']).center(WidthOfResult))
    if IsScanTotalEn=='enable':
        PassRate = format((float(LogResultDict['PassNum'])/float(TotalNum)),'.2%')
        StatisInfoList.append("   TOTAL PASSING Rate".ljust(WidthOfLogName) + ':' + "{}".format(PassRate).center(WidthOfResult))
    else:
        PassRate = format((float(LogResultDict['PassNum'])/float(ExecuteNum)),'.2%')
        StatisInfoList.append("   EXECUTED PASSING Rate".ljust(WidthOfLogName) + ':' + "{}".format(PassRate).center(WidthOfResult))
    StatisInfo = ""
    for statis_info in StatisInfoList:
        StatisInfo = StatisInfo + statis_info + '\n'
    '''==================================================================CovInfo================================================='''
    CovInfoList = []
    CovModeList = CovResultDict.keys()
    mode_idx = 0
    CovInfoHead = ""
    CovInfoHead = CovInfoHead + "|" + "MODE".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "code_coverage".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "LINE".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "COND".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "TOGL".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "FSM".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|" + "func_coverage".center(WidthOfResult)
    CovInfoHead = CovInfoHead + "|"
    if len(CovModeList)==0:
        CovInfoList.append("There is no any coverage get")
    else:
        CovInfoList.append(CovInfoHead)
        CovInfoList.append(SplitFlag1.replace('\n',''))
        for mode in CovModeList:
            tmp_context = ""
            tmp_context = tmp_context + "|" + "{}".format(mode).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][0]).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][1]).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][2]).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][3]).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][4]).center(WidthOfResult)
            tmp_context = tmp_context + "|" + "{}".format(CovResultDict[mode][5]).center(WidthOfResult)
            tmp_context = tmp_context + "|"
            CovInfoList.append(tmp_context)
            mode_idx = mode_idx + 1
            if mode_idx < len(CovModeList):
                CovInfoList.append(SplitFlag1.replace('\n',''))
    CovInfo = ""
    for cov_info in CovInfoList:
        CovInfo = CovInfo + cov_info + '\n'
    '''==================================================================ResultInfo================================================='''
    ResultInfoList =[]
    for log_result in LogResultDict['LogInfoList']:
        tmp_log = log_result[0]
        tmp_result = log_result[1]
        tmp_cputime = log_result[2]
        tmp_memory = log_result[3]
        tmp_lineinfo = log_result[4]
        ResultInfoList.append(tmp_log.ljust(WidthOfLogName) + '|' + tmp_result.center(WidthOfResult) +'|' + tmp_cputime.center(widthOfInfo) +'|' + tmp_memory.center(widthOfInfo) +'|'+ tmp_lineinfo)
    ResultInfo = ""
    for result_info in ResultInfoList:
        ResultInfo = ResultInfo + result_info + '\n'
    '''==================================================================CmdInfo================================================='''
    CmdInfo = ""
    CmdIdex = 0
    for cmd_info in CmdList:
        CmdInfo = CmdInfo + LogResultDict['LogInfoList'][CmdIdex][0] + " : " + cmd_info + "\n"
        CmdIdex = CmdIdex + 1
    HeadContext = SplitFlag0 + Title0 + SplitFlag1 + VersionInfo + SplitFlag1 + TimeInfo + SplitFlag1 + StatisInfo + SplitFlag1 + CovInfo + SplitFlag0
    logContext = HeadContext + Title1 + SplitFlag1 + ResultInfo + SplitFlag0 + Title2 + SplitFlag1 + CmdInfo + SplitFlag0
    '''==================================================================PrintInfo================================================='''
    if os.path.exists("{}/regress_result_dir".format(SimPath)) is False:
        os.makedirs("{}/regress_result_dir".format(SimPath))
    printEndInfo = "    More information in " + "{}/regress_result_dir/{}".format(SimPath,log_name) + '\n'
    PrintInfo = HeadContext + printEndInfo + SplitFlag0
    RegressLogFile = open("{}/regress_result_dir/{}".format(SimPath,log_name),"w")
    RegressLogFile.write(logContext)
    RegressLogFile.close()
    return PrintInfo,PassRate


def GenRegressResultExcelInfo(TotalNum, ExecuteNum, LogResultDict,CovResultDict, PassRate, ClosedTapdNum, NotClosedTapdNum,InfoDict):
    ExcelInfoList = []
    tmp_time = time.strftime("%Y%m%d_%H%M%S",time.localtime())
    for key,value in CovResultDict.items():
        ModeInfoResult = [tmp_time[0:8],TotalNum,ExecuteNum
                 ,LogResultDict['FatalNum'],LogResultDict['ErrorNum'],LogResultDict['WarningNum']
                 ,LogResultDict['RunningNum'],LogResultDict['NotStartNum'],LogResultDict['PassNum']
                 ,PassRate,key
                 ]
        ModeInfoResult.extend(value)
        ModeInfoResult.append(ClosedTapdNum)
        ModeInfoResult.append(NotClosedTapdNum)
        print(ModeInfoResult)
        ExcelInfoList.append(ModeInfoResult)
    return ExcelInfoList


def GenDB(InfoDict, TotalNum, ExecuteNum, LogResultDict, CovResultDict, PassRate):
    tmpDB = SqlMethod(InfoDict['CovInstanceName']+'_DB')
    EndRegressTime = time.strftime("%Y%m%d_%H%M%S",time.localtime())
    for ModeName in InfoDict['ModeList']:
        tmpDB.BuildNewTable(ModeName)
        Result =[
                tmpDB.GetLastId(ModeName)+1,                # ID
                EndRegressTime,                             # DATE
                TotalNum,                                   # TOTAL
                ExecuteNum,                                 # EXECUTED
                LogResultDict['FatalNum'],                  # FATAL
                LogResultDict['ErrorNum'],                  # ERROR
                LogResultDict['WarningNum'],                # WARNING
                LogResultDict['RunningNum'],                # RUNNING
                LogResultDict['NotStartNum'],               # NOT_START
                LogResultDict['PassNum'],                   # PASS
                PassRate,                                   # PASSRate
                CovResultDict[ModeName][0],                 # code_coverage
                CovResultDict[ModeName][1],                 # LINE
                CovResultDict[ModeName][2],                 # COND
                CovResultDict[ModeName][3],                 # TOGL
                CovResultDict[ModeName][4],                 # FSM
                CovResultDict[ModeName][5],                 # func_couverage
                InfoDict['ClosedTapdNum'],                  # Tapd_closed
                InfoDict['NotClosedTapdNum'],               # Tapd_not_closed
                InfoDict['PlanCovDict'][ModeName]           # PlanCoverage
                ]
        tmpDB.InsertValue(ModeName,Result)


def GenRegressResultWXInfo(TotalNum,LogResultDict, CovResultDict, PassRate):
    WXContent = ""
    if PassRate == "100.00%":
        WXContent += 'PassRate is 100.00% \n'
    else:
        WXContent += 'PassRate is {} :\n'.format(PassRate)
        if(LogResultDict['FatalNum'] != 0): WXContent += 'Fatal num is {}/{} \n'.format(LogResultDict['FatalNum'],TotalNum) 
        if(LogResultDict['ErrorNum'] != 0): WXContent += 'Error num is {}/{} \n'.format(LogResultDict['ErrorNum'],TotalNum)
        if(LogResultDict['WarningNum'] != 0): WXContent += 'Warning num is {}/{} \n'.format(LogResultDict['WarningNum'],TotalNum) 
        if(LogResultDict['RunningNum'] != 0): WXContent += 'Running num is {}/{} \n'.format(LogResultDict['RunningNum'],TotalNum) 
        if(LogResultDict['NotStartNum'] != 0):WXContent += 'Not Start num is {}/{} \n'.format(LogResultDict['NotStartNum'],TotalNum) 
    return WXContent


def gen_title_context(title_dict):
    title_context = '\t\t\t<tr>\n'
    for key in title_dict.keys():
        title_context += '\t\t\t\t<th{}>{}</th>\n'.format(title_dict[key],key)
    title_context += '\t\t\t</tr>'
    return title_context


def gen_table_context(context_list):
    table_context = ''
    for line in context_list:
        table_context += '\t\t\t<tr>\n'
        for col in line:
            table_context += '\t\t\t\t<th>{}</th>\n'.format(col)
        table_context += '\t\t\t</tr>\n'
    return table_context


def gen_table(title_context,table_context):
    context = '''
        <table border="1" cellspacing="0">
{_title_context}
{_table_context}
        </table>
'''.format(_title_context=title_context,_table_context=table_context)
    return context


def GenRegressResultEmail(InfoDict,LogResultDict,CovResultDict,PassRate,ExecuteNum):
    Content = ''
    # Regress Time:
    EndRegressTime = time.strftime("%Y%m%d_%H%M%S",time.localtime())
    tmpdict = {'Regress Time Info':' bgcolor="#CDCDCD" colspan="2"'}
    tmpTitle = gen_title_context(tmpdict)
    tmpdict = {'StartRegressTime':' bgcolor="#CDCDCD"','EndRegressTime':' bgcolor="#CDCDCD"'}
    tmpTitle += gen_title_context(tmpdict)
    tmpContext = gen_table_context([[InfoDict['StartRegressTime'],EndRegressTime]])
    Content += gen_table(tmpTitle,tmpContext)
    Content +='<p></p>'
    # Max CPU Cost:
    tmpdict = {'CPU Cost Info':' bgcolor="#CDCDCD" colspan="2"'}
    tmpTitle = gen_title_context(tmpdict)
    tmpdict = {'Max CPU Time':' bgcolor="#CDCDCD"','Max CPU Memory':' bgcolor="#CDCDCD"'}
    tmpTitle += gen_title_context(tmpdict)
    tmpContext = gen_table_context([[LogResultDict['MaxCpuTimeInfo'][2],LogResultDict['MaxMemoryInfo'][3]]])
    Content += gen_table(tmpTitle,tmpContext)
    Content +='<p></p>'
    # Result
    tmpdict = {'Regress Result':' bgcolor="#CDCDCD" colspan="8"'}
    tmpTitle = gen_title_context(tmpdict)
    tmpdict = {'PassRate':' bgcolor="#CDCDCD"',
            'EXECUTED':' bgcolor="#CDCDCD"',
            'FATAL':' bgcolor="#CDCDCD"',
            'ERROR':' bgcolor="#CDCDCD"',
            'WARNING':' bgcolor="#CDCDCD"',
            'RUNNING':' bgcolor="#CDCDCD"',
            'NOT START':' bgcolor="#CDCDCD"',
            'PASS':' bgcolor="#CDCDCD"'
            }
    tmpTitle += gen_title_context(tmpdict)
    tmpContext = gen_table_context([[
        PassRate,
        ExecuteNum,
        LogResultDict['FatalNum'],
        LogResultDict['ErrorNum'],
        LogResultDict['WarningNum'],
        LogResultDict['RunningNum'],
        LogResultDict['NotStartNum'],
        LogResultDict['PassNum']
        ]])
    Content += gen_table(tmpTitle,tmpContext)
    Content +='<p></p>'
    # Coverage:
    tmpdict = {'Coverage Info':' bgcolor="#CDCDCD" colspan="8"'}
    tmpTitle = gen_title_context(tmpdict)
    tmpdict = {'Mode':' bgcolor="#CDCDCD"',
            'Code Coverage':' bgcolor="#CDCDCD"',
            'Line Coverage':' bgcolor="#CDCDCD"',
            'Condition Coverage':' bgcolor="#CDCDCD"',
            'Toggle Coverage':' bgcolor="#CDCDCD"',
            'FSM Coverage':' bgcolor="#CDCDCD"',
            'Function Coverage':' bgcolor="#CDCDCD"',
            'Plan Coverage':' bgcolor="#CDCDCD"',
            }
    tmpTitle += gen_title_context(tmpdict)
    tmpList = []
    CovModeList = CovResultDict.keys()
    if len(CovModeList)==0:
        CovInfoList.append("There is no any coverage get")
    else:
        for mode in CovModeList:
            tmpList.append([mode]+CovResultDict[mode]+[InfoDict['PlanCovDict'][mode]])
    tmpContext = gen_table_context(tmpList)
    Content += gen_table(tmpTitle,tmpContext)
    Content += '<p></p> \n'
    return Content


def GenRegressResult(SimPath,IsScanTotalEn,TotalNum,ExecuteNum,InfoDict,LogResultDict,CovResultDict,ProjectVersion,CmdList):
    PrintInfo,PassRate = GenRegressResultLog(SimPath,IsScanTotalEn,TotalNum,ExecuteNum,InfoDict,LogResultDict,CovResultDict,ProjectVersion,CmdList)
    if IsScanTotalEn == "enable":
        WXContent = GenRegressResultWXInfo(TotalNum,LogResultDict,CovResultDict,PassRate)
    else:
        WXContent = GenRegressResultWXInfo(ExecuteNum,LogResultDict,CovResultDict,PassRate)
    EmailInfo = GenRegressResultEmail(InfoDict,LogResultDict,CovResultDict,PassRate,ExecuteNum)
    ExcelInfoList = GenRegressResultExcelInfo(TotalNum,ExecuteNum,LogResultDict,CovResultDict,PassRate,InfoDict['ClosedTapdNum'],InfoDict['NotClosedTapdNum'],InfoDict)
    GenDB(InfoDict,TotalNum,ExecuteNum,LogResultDict,CovResultDict,PassRate)
    return ExcelInfoList,PrintInfo,WXContent,EmailInfo


def send_to_host(ExcelFilePath,ExcelName):
    # Please, change ip, port and password to your own infos;
    # eg ip = "1.234.123.111"
    #    port = "36000"
    #    password = "password"
    LocalFile = ExcelFilePath
    RemotePath = '/data/regress_result'
    RemoteFile = RemotePath +'/' + ExcelName + '.xlsx'
    ip = ""
    port = ""
    user = ""
    password = ""
    try:
        ssh_scp_put(ip,port,user,password,LocalFile,RemoteFile)
    except:
        print('Remote Transportation Error!!!')


def ssh_scp_put(ip, port, user, password, local_file, remote_file):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, port, user, password)
    sftp = ssh.open_sftp()
    sftp.put(local_file, remote_file)
    sftp.close()


def GenRegressExcel(ExcelPath,ExcelName):
    ExcelFileName = ExcelName+'.'+'xlsx'
    ExcelFilePath = os.path.abspath(os.path.join(ExcelPath,ExcelFileName))
    path = pathlib.Path(ExcelFilePath)
    if path.is_file():
        print('ExcelFile dir',ExcelFilePath)
    elif ExcelName == '':
        ExcelFilePath = ''
        print('no ExcelFile created due to lack of excel_name declare')
    else:
        value_title = ['DATE','TOTAL','EXECUTED','FATAL','ERROR','WARNING','RUNNING','NOT START','PASS','PASS Rate','Mode','code_coverage','LINE','COND','TOGL','FSM','func_coverage','Tapd_closed','Tapd_not_closed']
        RegressExcel = xlwt.Workbook()
        sheet = RegressExcel.add_sheet(ExcelName)
        for i in range(0, len(value_title)):
            sheet.write(0,i,value_title[i])
        RegressExcel.save(ExcelFilePath)
        print('ExcelFilePath',ExcelFilePath)
    return ExcelFilePath


def AppendRegressExcel(ExcelFilePath,value):
    if ExcelFilePath:
        workbook = xlrd.open_workbook(ExcelFilePath,formatting_info=True)
        sheets = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheets[0])
        rows_old = worksheet.nrows
        cols_num = worksheet.ncols
        new_workbook = xlcopy(workbook)
        last_date = worksheet.cell_value((rows_old-1),0)
        new_worksheet = new_workbook.get_sheet(0)
        for col in range(cols_num):
            if col > 8 and col != 10 and col < 17:
               style_percent = xlwt.easyxf(num_format_str='0.00%')
            else:
               style_percent = xlwt.easyxf(num_format_str='General')
            for row in range(len(value)):
                ''' override same date datas '''
                if row == 0 and col == 0:
                    if last_date == value[0][0]:
                        start_row = rows_old -1
                    else:
                        start_row = rows_old
                ''' write data '''
                if row == 0 or col > 9:
                    new_worksheet.write(start_row + row, col, value[row][col],style_percent)
        print('regress result of ',value[0][0],'has been recorded')
        new_workbook.save(ExcelFilePath)


def DoExcelWrite(ExcelInfoList,DoTransport,ExcelName,ExcelPath,DoTransWhileRegr,GetWhileRegrFlag):
    ExcelFilePath = GenRegressExcel(ExcelPath,ExcelName)
    AppendRegressExcel(ExcelFilePath,ExcelInfoList)
    if (GetWhileRegrFlag == 'disable') and (DoTransport == "enable"):
        send_to_host(ExcelFilePath,ExcelName)
    elif (GetWhileRegrFlag == 'enable') and (DoTransWhileRegr == 'enable') and (DoTransport == "enable"):
        send_to_host(ExcelFilePath,ExcelName)
    else:
        print()


def GenErrorIni(SimPath,InfoDict,CmdList,LogResultDict):
    ''' get error cmd '''
    LogInfoList = LogResultDict['LogInfoList']
    ErrorCmd = []
    for i in range(len(LogInfoList)):
        LogInfo = LogInfoList[i]
        if LogInfo[1] == "Fatal" or LogInfo[1] == "Error":
            ErrorCmd.append(CmdList[i])
    ''' generate error tc_list'''
    if len(ErrorCmd) == 0:
        return
    ErrTcList= []
    for Cmd in ErrorCmd:
        ErrTcList.append(GetOneCmdString(InfoDict,Cmd))
    ''' goto regress path '''
    RegressPath = SimPath + "/../regress"
    os.chdir(RegressPath)
    ''' generate new ini file'''
    if os.path.exists("{}/error_ini_dir".format(SimPath)) is False:
        os.makedirs("{}/error_ini_dir".format(SimPath))
    errIniFile = "{}/error_ini_dir/error_{}.ini".format(SimPath,InfoDict['StartRegressTime'])
    GenNewIni(InfoDict,ErrTcList,SimPath,errIniFile,'enable')


def GenNewIni(InfoDict,ErrTcList,SimPath,IniName,CmpFlag):
    errIni = ConfigParser()
    errIni['GENERAL'] = {
            'tools':InfoDict['ToolInfo'],
            'operation':'none',
            'url':'',
            'clone_workspace':'',
            'sim_path':SimPath,
            'module_name':InfoDict['CovInstanceName'],
            'excel_name':InfoDict['ExcelName'],
            'excel_path':InfoDict['ExcelPath'],
            'sim_tools':InfoDict['SimTools'],
            'paral_run_num':6,
            'anti_annotation_enable':InfoDict['AntiAnnotationEnable'],
            'testpoint_file_name':InfoDict['TestpointFileName'],
            'fail_recmp':'disable',
            'recmp_opts':'',
            'fail_rerun':'disable',
            'rerun_opts':'',
            'ignore_warning_en'       : 'disable',
            'fatal_key_words         ': '[\"'+'\",\"'.join(InfoDict['KewWordDict']['FatalKeyWordList']) + '\"]',
            'error_key_words         ': '[\"'+'\",\"'.join(InfoDict['KewWordDict']['ErrorKeyWordList']) + '\"]',
            'warning_key_words       ': '[\"'+'\",\"'.join(InfoDict['KewWordDict']['WarningKeyWordList']) + '\"]',
            'ignore_fatal_key_words  ': str(InfoDict['KewWordDict']['IgnoreFatalDict']),
            'ignore_error_key_words  ': str(InfoDict['KewWordDict']['IgnoreErrorDict']),
            'ignore_warning_key_words': str(InfoDict['KewWordDict']['IgnoreWarningDict']),
            'total_ignore_log_list   ': '[\"'+'\",\"'.join(InfoDict['TotalIgnoreLogList']) + '\"]',
            'start_chk_fatal_key_word': InfoDict['StartChkFatalWord'],
            'start_chk_error_key_word': InfoDict['StartChkErrorWord'],
            'start_chk_warning_key_word': InfoDict['StartChkWarningWord'],
            'set_daily_regress':'disable',
            'regress_date_hour':'0',
            'regress_date_minute':'0',
            'project_source_file':''
            }
    errIni['PRE_REGRESSION'] = {
            'pre_script_enable':'disable',
            'pre_make_enable'  :'disable',
            'post_script_enable' :'disbale',
            'post_make_enable': 'disbale',
            'pre_script' :'',
            'pre_make':'',
            'post_script':'',
            'post_make':''
            }
    errIni['REGRESSION'] = {
            'dotransport': 'disable',
            'transport_while_regress':'disable',
            'docompile'  : CmpFlag,
            'fcov':InfoDict['FcovEn'],
            'ccov':InfoDict['CcovEn'],
            'fcov_key_word':InfoDict['FcovKeyWord'],
            'udf':'',
            'cmp_opts':InfoDict['CmpOptions'],
            'dorotation':InfoDict['DoRotation'],
            'del_pass_log':InfoDict['DelPassLog'],
            'del_err_cov':InfoDict['DelErrCov'],
            'tc_list':'[\n' + ''.join(ErrTcList) + ']'
            }
    if os.path.exists(IniName):
        os.remove(IniName)
    os.mknod(IniName)
    with open(IniName,'w') as file:
        errIni.write(file)

def to_percent(temp,position):
    return "%.2f%%"%(100*float(temp))


def p2f(x):
    return float(x.strip('%'))/100


def f2p(x):
    return "{:.2%}".format(x)


def KeyDate2Mode(CovDict):
    NewCovDict = {}
    for date in CovDict.keys():
        for ModeCov in CovDict[date]:
            DateDic = {}
            tmpMode = ModeCov['Mode']
            del ModeCov['Mode']
            DateDic[date] = ModeCov
            if tmpMode not in NewCovDict.keys():
                NewCovDict[tmpMode] = DateDic
            else:
                NewCovDict[tmpMode][date] = ModeCov
    return NewCovDict


def GenCovXYList(Idx,CovDict):
    CovName = ['CCov','Line','Cond','TGL','FSM','FCov']
    xlist = []
    ylist = []
    for date in CovDict.keys():
        if CovDict[date][CovName[Idx]] != '--':
            xlist.append(date)
            ylist.append(p2f(CovDict[date][CovName[Idx]]))
    xlist.reverse()
    ylist.reverse()
    return xlist, ylist


def GenTapdXYList(Idx,TapdDict):
    xlist = []
    ylist = []
    for date in TapdDict.keys():
        xlist.append(date)
        ylist.append(int(TapdDict[date][Idx]))
    xlist.reverse()
    ylist.reverse()
    return xlist, ylist


def GenPassImg(InfoDict,PassImgName,PassRateDict,DaysDelta):
    DaysList = []
    for i in range(DaysDelta-1,-1,-1):
        tmpDate = str(datetime.date.today() - datetime.timedelta(days=i))
        tmpDateStr = tmpDate.split('-')[0] + tmpDate.split('-')[1] + tmpDate.split('-')[2]
        DaysList.append(tmpDateStr)
    plt.figure(figsize=(12,7))
    plt.xlabel('Time',fontsize=10)
    #plt.xticks(range(len(DaysList)),DaysList,rotation=90,fontsize=12)
    plt.grid()
    plt.gca().yaxis.set_major_formatter(FuncFormatter(to_percent))
    xlist = list(PassRateDict.keys())
    xlist.reverse()
    ylist = list(PassRateDict.values())
    ylist.reverse()
    plt.plot_date(xlist,ylist,linestyle='solid',marker="o",label="Pass Rate")
    plt.title(PassImgName,fontsize=20,fontweight='heavy',loc='center')
    plt.legend(loc='best',fontsize=15,bbox_to_anchor=(1,0,0.2,1))
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig('./'+ PassImgName +'.png')
    return


def GenCovImg(InfoDict,CovImgName,CovDict,DaysDelta):
    DaysList = []
    LabelNameList= ['Code Cov', 'Line Cov', 'Cond Cov', 'Togl Cov', 'FSM Cov','Func Cov']
    for i in range(DaysDelta-1,-1,-1):
        tmpDate = str(datetime.date.today() - datetime.timedelta(days=i))
        tmpDateStr = tmpDate.split('-')[0] + tmpDate.split('-')[1] + tmpDate.split('-')[2]
        DaysList.append(tmpDateStr)
    plt.figure(figsize=(12,7))
    plt.xlabel('Time',fontsize=10)
    # plt.xticks(range(len(DaysList)),DaysList,rotation=90,fontsize=12)
    plt.grid()
    plt.gca().yaxis.set_major_formatter(FuncFormatter(to_percent))
    for i in range(6):
        xlist,ylist = GenCovXYList(i,CovDict)
        if xlist != []:
            plt.plot_date(xlist,ylist,linestyle='solid',marker="o",label=LabelNameList[i])
        else:
            plt.plot_date([str(datetime.date.today())],["0%"],linestyle='solid',marker="o",label=LabelNameList[i])
    plt.title(CovImgName,fontsize=20,fontweight='heavy',loc='center')
    plt.legend(loc='best',fontsize=15,bbox_to_anchor=(1,0,0.2,1))
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig('./'+ CovImgName +'.png')
    return


def GenTapdImg(InfoDict,TapdImgName,TapdDict,DaysDelta):
    DaysList = []
    LabelNameList= ['Tapd Closed','Tapd Not Closed']
    for i in range(DaysDelta-1,-1,-1):
        tmpDate = str(datetime.date.today() - datetime.timedelta(days=i))
        tmpDateStr = tmpDate.split('-')[0] + tmpDate.split('-')[1] + tmpDate.split('-')[2]
        DaysList.append(tmpDateStr)
    plt.figure(figsize=(12,7))
    plt.xlabel('Time',fontsize=10)
    plt.grid()
    for i in range(2):
        xlist,ylist = GenTapdXYList(i,TapdDict)
        if xlist != []:
            plt.plot_date(xlist,ylist,linestyle='solid',marker="o",label=LabelNameList[i])
        else:
            plt.plot_date([str(datetime.date.today())],[0],linestyle='solid',marker="o",label=LabelNameList[i])
    plt.title(TapdImgName,fontsize=20,fontweight='heavy',loc='center')
    plt.legend(loc='best',fontsize=15,bbox_to_anchor=(1,0,0.2,1))
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig('./'+ TapdImgName +'.png')
    return

def GetModuleClosedNum(client,TargetModule,TargetStatus):
    ClosedNum = 0
    NotClosedNum = 0
    TapdNumPerPage = 200
    tmpDic = client.get('/bugs/count')
    TotalTapdNum = tmpDic['data']['count']
    PageNum = TotalTapdNum // TapdNumPerPage if (TotalTapdNum % TapdNumPerPage == 0) else TotalTapdNum // TapdNumPerPage + 1
    # print('total Tapd Num is {}, in per page limit is {}, page num is {}'.format(TotalTapdNum,TapdNumPerPage,PageNum))
    for pageIdx in range(1,PageNum+1):
        tmpDic = client.get('/bugs',{'limit':TapdNumPerPage,'page':pageIdx})
        for tapd_data in tmpDic['data']:
            if tapd_data['Bug']['module'] == TargetModule and tapd_data['Bug']['status'] == TargetStatus:
                ClosedNum += 1
            elif tapd_data['Bug']['module'] == TargetModule and tapd_data['Bug']['status'] != TargetStatus:
                NotClosedNum += 1
                print(tapd_data['Bug']['module'],'  ',tapd_data['Bug']['status'],'   ',tapd_data['Bug']['created'])
        # print('In page {}, total {}'.format(pageIdx,len(tmpDic['data'])))
    return ClosedNum, NotClosedNum


def GetTapdInfo(TapdEnable,TapdModuleName):
    if TapdEnable == 'enable':
        # please sign up your own api user in http://http://o.tapd.oa.com
        TapdTestProject = ""
        TapdTestPw = ""
        client = TapdClient(TapdTestProject, TapdTestPw)
        # print(client.authentication()) # test Tapd api link
        TargetModule = TapdModuleName
        TargetStatus = 'closed'
        ClosedTapdNum,NotClosedTapdNum = GetModuleClosedNum(client,TargetModule,TargetStatus)
    else:
        ClosedTapdNum = 0
        NotClosedTapdNum = 0
    return ClosedTapdNum, NotClosedTapdNum


def DoAntiAnnotation(AntiAnnotationEnable,TestpointFileName,InfoDict):
    PlanCovDict = {}
    for ModeName in InfoDict['ModeList']:
        PlanCovDict[ModeName] = '--'
    TestpointFilePath = './'
    if AntiAnnotationEnable == 'enable' and InfoDict['SimTools'] == 'xrun':
        for ModeName in InfoDict['ModeList']:
            tmpFcovClass = XrunTestPointReverse()
            FcovDict = tmpFcovClass.Mode2FcovDict(ModeName)
            a = anti_annotation(FcovDict, TestpointFilePath, TestpointFileName)
            PlanCovDict[ModeName] = f2p(a.WriteAntiAnnotation())
        # generate sum cov result
        os.system('make cov_txt_sum')
        tmpFcovClass = XrunTestPointReverse()
        FcovDict = tmpFcovClass.Mode2FcovDict('')
        a = anti_annotation(FcovDict, TestpointFilePath, TestpointFileName)
        TotalPlanCov = f2p(a.WriteAntiAnnotation())
    return PlanCovDict


def GenRegressLog(SimPath,IsScanTotalEn,CovFlag='enable',GetWhileRegrFlag='disable',GetNotDelPassLog="disable"):
    os.chdir(SimPath)
    if not os.path.exists("./cmd.file"):
        print("there is no cmd.file for regress_log analysis",flush = True)
        sys.exit()
    # calculate log info
    InfoDict = GetRegressCfgInfo(SimPath)
    TotalLogList = GetTotalLogList(SimPath,InfoDict['TotalIgnoreLogList'],InfoDict['ModeList'])
    if IsScanTotalEn=='enable':
        LogList = TotalLogList
        CmdList = []
    else:
        LogList = InfoDict['ExecuteLogList']
        CmdList = InfoDict['ExecuteCmdList']
    if GetNotDelPassLog=='enable':
        LogResultDict = DoLogScan(LogList,InfoDict['KewWordDict'],'disable',[InfoDict['StartChkFatalWord'],InfoDict['StartChkErrorWord'],InfoDict['StartChkWarningWord']],InfoDict['SimTools'])
    else:
        LogResultDict = DoLogScan(LogList,InfoDict['KewWordDict'],InfoDict['DelPassLog'],[InfoDict['StartChkFatalWord'],InfoDict['StartChkErrorWord'],InfoDict['StartChkWarningWord']],InfoDict['SimTools'])
    if CovFlag == 'enable':
        CovResultDict = GetCovResult(SimPath,InfoDict['ModeList'],InfoDict['CcovEn'],InfoDict['FcovEn'],InfoDict['CovInstanceName'],InfoDict['SimTools'],InfoDict)
    else:
        CovResultDict = {}
        for Mode in InfoDict['ModeList']:
            CovResult = ['--','--','--','--','--','--']
            CovResultDict[Mode] = CovResult
    # generate regress_result info
    ProjectVersion = GetProjectVersion(SimPath,InfoDict['ToolInfo'])
    InfoDict['ClosedTapdNum'], InfoDict['NotClosedTapdNum'] = GetTapdInfo(InfoDict['TapdEnable'],InfoDict['TapdModuleName'])
    InfoDict['PlanCovDict'] =DoAntiAnnotation(InfoDict['AntiAnnotationEnable'],InfoDict['TestpointFileName'],InfoDict)
    TotalNum = len(TotalLogList)
    ExecuteNum = len(LogList)
    ExcelInfoList,PrintInfo,WXContent,EmailInfo = GenRegressResult(SimPath,IsScanTotalEn,TotalNum,ExecuteNum,InfoDict,LogResultDict,CovResultDict,ProjectVersion,CmdList)
    DoExcelWrite(ExcelInfoList,InfoDict['DoTransPort'],InfoDict['ExcelName'],InfoDict['ExcelPath'],InfoDict['TransportWhileRegress'],GetWhileRegrFlag)
    print(PrintInfo)
    if IsScanTotalEn=='disable':
        GenErrorIni(SimPath,InfoDict,CmdList,LogResultDict)


def GetRegressStatus():
    try:
        GetRegressStatusEN = sys.argv[1]
    except:
        GetRegressStatusEN = "disable"
    tmp_IsScanTotalEn = 'disable'
    tmp_GetCovEn = 'enable'
    tmp_GetWhileRegrFlag = 'disable'
    tmp_GetNotDelPassLog = "disable"
    if GetRegressStatusEN == "regress_status":
        tmp_IsScanTotalEn = 'disable'
        tmp_GetCovEn = 'disable'
        tmp_GetWhileRegrFlag = 'enable'
        tmp_GetNotDelPassLog = "enable"
    elif GetRegressStatusEN == "regress_status_cov":
        tmp_IsScanTotalEn = 'disable'
        tmp_GetCovEn = 'enable'
        tmp_GetWhileRegrFlag = 'enable'
        tmp_GetNotDelPassLog = "enable"
    elif GetRegressStatusEN == "total_status":
        tmp_IsScanTotalEn = 'enable'
        tmp_GetCovEn = 'enable'
        tmp_GetWhileRegrFlag = 'enable'
        tmp_GetNotDelPassLog = 'enable'
    else:
        return
    CurrentPath = './'
    tmp_SimPath = os.path.abspath(os.path.join(CurrentPath,'../sim'))
    GenRegressLog(tmp_SimPath,tmp_IsScanTotalEn,tmp_GetCovEn,tmp_GetWhileRegrFlag,tmp_GetNotDelPassLog)
    sys.exit()


def GetRegressTime():
    try:
        CurrTime = sys.argv[2]
    except:
        CurrTime = time.strftime("%Y%m%d_%H%M%S",time.localtime())
    return CurrTime


if __name__=="__main__":
    RegressTime = GetRegressTime()
    GetRegressStatus()
    tmp_GeneralDict,tmp_PreRegressDict,tmp_RegressDict = DoIniParser()
    tmp_SimPath = GenRegressPath(tmp_GeneralDict)
    DoPreRegress(tmp_SimPath,tmp_PreRegressDict)
    DoRegress(tmp_SimPath,tmp_GeneralDict,tmp_PreRegressDict,tmp_RegressDict)
    DoPostRegress(tmp_SimPath,tmp_PreRegressDict)
    tmp_IsScanTotalEn = 'disable'
    GenRegressLog(tmp_SimPath,tmp_IsScanTotalEn)

