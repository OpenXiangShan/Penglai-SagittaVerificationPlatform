#!/usr/bin/python3
# coding=utf-8
import os
import subprocess
import time
import sys
from openpyxl import load_workbook
from numpy import *

# pip install openpyxl

class ExcelOp(object):

    sheet = None

    def __init__(self,file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet_all = sheets
        self.sheet = sheets[0]
        #self.ws = self.wb[self.sheet]

    def get_row_clo_num(self):
        self.ws = self.wb[self.sheet]
        rows = self.ws.max_row
        colums = self.ws.max_column
        return rows,columns

    def get_cell_value(self,row,column):
        self.ws = self.wb[self.sheet]
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    def get_row_value(self,row):
        self.ws = self.wb[self.sheet]
        columns = self.ws.max_column
        row_data = []
        for i in range(1,columns + 1):
            cell_value = self.ws.cell(row=row,column=i).value
            row_data.append(cell_value)
        return row_data

    def get_col_value(self,column):
        self.ws = self.wb[self.sheet]
        rows = self.ws.max_row
        column_data = []
        for i in range(1,rows + 1):
            cell_value = self.ws.cell(row=i,column=column).value
            column_data.append(cell_value)
        return column_data

    def set_cell_value(self,row,column,cellvalue):
        try:
            self.ws = self.wb[self.sheet]
            self.ws.cell(row=row,column=column).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws = self.wb[self.sheet]
            self.ws.cell(row=row,column=column).value = "writefail"
            self.wb.save(self.file)



def coverage_revert(tar_file):
   
    ##get file:urgReport/groups.txt
    urg_groups_file = open("./urgReport/groups.txt")
    urg_lines = urg_groups_file.readlines()
    #print(urg_lines)


    excel_op = ExcelOp(tar_file)
    for sheet_index in excel_op.sheet_all :
        print('sheet(name:' + sheet_index +  ') exits in xlsx!! Paring!!')
        excel_op.sheet = sheet_index
        first_row = excel_op.get_row_value(1)
        #print(first_row)
        if 'reverse_testpoint' in first_row and 'weight' in first_row and 'reverse_percent' in first_row:
            print('Get KEY: reverse_testpoint/weight/reverse_percent sucessfully !')
            #print(excel_op.get_col_value(first_row.index('reverse_testpoint')+1))
        else:
            print('Get KEY: reverse_testpoint/weight/reverse_percent fail !  Ending!')
            continue
        
        #get all test_point_list
        test_point_alllist = list((excel_op.get_col_value(first_row.index('reverse_testpoint')+1)))[1:]
        #print(test_point_alllist)
        
        
        test_point_list = None
        row_index = 0
        percent_all = []
        weight_list = []
        for test_point_in_one_cell in test_point_alllist:
            if test_point_in_one_cell is not None  :
                #print(test_point_in_one_cell)
                row_index = test_point_alllist.index(test_point_in_one_cell)
                #print('index=' + str(row_index))
                test_point_list = test_point_in_one_cell.split() #get one cell's test_points
            else:
                continue
            #get testpoint's list in one cell
            #print('------------------')
            print(test_point_list)
            percent_list = []
            for test_point in test_point_list:
                #print(test_point)
                for urg_line in urg_lines:
                    if test_point in urg_line:
                        #print(urg_line.split())
                        percent_list.append(double(urg_line.split()[0]))
                    else:
                        continue
            #print(mean(percent_list))       
    
            excel_op.set_cell_value(row=row_index+2,column=first_row.index('reverse_percent')+1,cellvalue=str(mean(percent_list))+'%')
            weight = double(excel_op.get_cell_value(row=row_index+2,column=first_row.index('reverse_percent')))
            weight_list.append(weight)
            #print('weight = '+str(excel_op.get_cell_value(row=row_index+2,column=first_row.index('reverse_percent'))))
            percent_all.append(mean(percent_list)*weight)
        
        excel_op.set_cell_value(row=2,column=first_row.index('reverse_percent')+1,cellvalue=str(sum(percent_all)/sum(weight_list))+'%')

            

if __name__=="__main__":
    coverage_revert(sys.argv[1])

